#!/usr/bin/env python3
"""
ADCS Web Audit Server v1.0
──────────────────────────────────────────────────────────────
Runs the ADCS audit, generates a rich HTML dashboard,
and serves it on http://0.0.0.0:4000

Usage:
  python3 adcs_web.py -u USER -p PASS -d DC_IP --domain FQDN
  Then open http://localhost:4000 in your browser.

Requires: pip install ldap3
"""

import argparse, sys, struct, json, html, os, hashlib, datetime, io, zipfile
from http.server import HTTPServer, BaseHTTPRequestHandler
from ldap3 import Server, Connection, ALL, NTLM, SIMPLE, SUBTREE
from ldap3.core.exceptions import LDAPBindError
from ldap3.protocol.microsoft import security_descriptor_control

SD_CONTROL = security_descriptor_control(sdflags=0x07)

# ══════════════════════════════════════════════════════════
# SD BINARY PARSER
# ══════════════════════════════════════════════════════════

ACE_ALLOWED, ACE_DENIED, ACE_ALLOWED_OBJ, ACE_DENIED_OBJ = 0,1,5,6

# ── GUID helpers ──────────────────────────────────────────────────────────
def _guid(data, off):
    """Parse a Windows GUID (mixed-endian) at offset, return lowercase string."""
    if off+16 > len(data): return None
    b = data[off:off+16]
    suffix = ''.join('{:02x}'.format(x) for x in b[10:16])
    return (
        '{:02x}{:02x}{:02x}{:02x}'.format(b[3],b[2],b[1],b[0]) + '-' +
        '{:02x}{:02x}'.format(b[5],b[4]) + '-' +
        '{:02x}{:02x}'.format(b[7],b[6]) + '-' +
        '{:02x}{:02x}'.format(b[8],b[9]) + '-' + suffix
    )

# Attribute GUIDs that directly affect certificate issuance security.
# WriteProperty scoped to any of these = exploitable for ESC4.
# WriteProperty scoped to any OTHER GUID = benign (false positive).
# ─── GUID classification ────────────────────────────────────────────────────
# These two GUIDs are ENROLLMENT EXTENDED RIGHTS (mask=DS_CTL_ACCESS/0x100).
# They mean "can request a certificate" — completely normal for Domain Users.
# They are NOT WriteProperty on template attributes and do NOT enable ESC4.
ENROLL_RIGHT_GUIDS = {
    "0e10c968-78fb-11d2-90d4-00c04f79dc55",  # Certificate-Enrollment  (Enroll)
    "a05b8cc2-17bc-4802-a710-e7c15ab866a2",  # Certificate-AutoEnrollment
}

# These GUIDs are actual msPKI template attributes.
# WriteProperty (mask=DS_WRITE_PROP/0x20) scoped to any of these = ESC4.
DANGEROUS_ATTR_GUIDS = {
    "d15ef7d8-f226-46db-ae79-b34e560bd12c",  # msPKI-Certificate-Name-Flag
    "ea1dddc4-60ff-416e-8cc0-17cee534bce7",  # msPKI-Enrollment-Flag
    "3f17c5c0-8f3e-4a4a-8e60-5c51ff1a0d64",  # msPKI-RA-Signature
    "1f298a89-de98-47b8-b5cd-572ad53d267e",  # msPKI-Certificate-Application-Policy
    "e5209ca2-3bba-11d2-90cc-00c04fd91ab1",  # pKIExtendedKeyUsage
    "4125c71f-7fac-4ff0-bcb7-f09a41325286",  # msPKI-Template-Schema-Version
    # Unscoped sentinel: standard ACE (no GUID) = all attributes
    "NO_GUID",
}

GUID_NAMES = {
    "d15ef7d8-f226-46db-ae79-b34e560bd12c": "msPKI-Certificate-Name-Flag",
    "ea1dddc4-60ff-416e-8cc0-17cee534bce7": "msPKI-Enrollment-Flag",
    "3f17c5c0-8f3e-4a4a-8e60-5c51ff1a0d64": "msPKI-RA-Signature",
    "1f298a89-de98-47b8-b5cd-572ad53d267e": "msPKI-Certificate-Application-Policy",
    "e5209ca2-3bba-11d2-90cc-00c04fd91ab1": "pKIExtendedKeyUsage",
    "4125c71f-7fac-4ff0-bcb7-f09a41325286": "msPKI-Template-Schema-Version",
    "0e10c968-78fb-11d2-90d4-00c04f79dc55": "Certificate-Enrollment",
    "a05b8cc2-17bc-4802-a710-e7c15ab866a2": "Certificate-AutoEnrollment",
    "bf967950-0de6-11d0-a285-00aa003049e2": "description",
    "f0f8ff84-1191-11d0-a060-00aa006c33ed": "userCertificate",
    "5b47d60f-6090-40b2-9f37-2a4de88f3063": "msDS-KeyCredentialLink",
    "e48d0154-bcf8-11d1-8702-00c04fb96050": "publicKeyPolicy",
    "NO_GUID": "(all attributes)",
}

def _sid(data, off):
    if off+8 > len(data): return "S-?"
    rev=data[off]; sub=data[off+1]
    auth=int.from_bytes(data[off+2:off+8],'big')
    subs=[str(struct.unpack_from('<I',data,off+8+i*4)[0])
          for i in range(sub) if off+8+i*4+4<=len(data)]
    return f"S-{rev}-{auth}"+"".join(f"-{s}" for s in subs)

def _dacl(data, off):
    """
    Parse DACL. For Object ACEs, preserve the ObjectType GUID so callers
    can determine whether WriteProperty is scoped to a dangerous attribute.
    Returns list of:
      { sid, mask, deny, object_guid (str|None), scoped (bool) }
    """
    aces=[]
    if off==0 or off+8>len(data): return aces
    count=struct.unpack_from('<H',data,off+4)[0]
    pos=off+8
    for _ in range(count):
        if pos+4>len(data): break
        atype=data[pos]; asize=struct.unpack_from('<H',data,pos+2)[0]
        deny=atype in(ACE_DENIED,ACE_DENIED_OBJ)

        if atype in(ACE_ALLOWED,ACE_DENIED):
            # Standard ACE — unscoped, affects ALL attributes
            if pos+8<=len(data):
                mask=struct.unpack_from('<I',data,pos+4)[0]
                aces.append({"sid":_sid(data,pos+8),"mask":mask,"deny":deny,
                             "object_guid":None,"scoped":False})

        elif atype in(ACE_ALLOWED_OBJ,ACE_DENIED_OBJ):
            # Object ACE — may be scoped to a specific attribute GUID
            if pos+12<=len(data):
                mask    =struct.unpack_from('<I',data,pos+4)[0]
                fl      =struct.unpack_from('<I',data,pos+8)[0]
                sid_off =pos+12
                obj_guid=None
                if fl&0x01:              # ObjectType present
                    obj_guid=_guid(data,sid_off)
                    sid_off+=16
                if fl&0x02:              # InheritedObjectType present
                    sid_off+=16
                if sid_off<=pos+asize:
                    aces.append({"sid":_sid(data,sid_off),"mask":mask,"deny":deny,
                                 "object_guid":obj_guid,"scoped":True})
        pos+=max(asize,4)
    return aces

def parse_sd(sb):
    """
    Parse SD. Returns per-ACE list (NOT merged) so callers can inspect GUIDs.
    Each entry: { sid, mask, deny, object_guid, scoped,
                  dangerous_write (bool), write_reason (str) }
    """
    if not sb or len(sb)<20: return []
    d=bytes(sb); doff=struct.unpack_from('<I',d,16)[0]
    if doff==0: return []

    aces=_dacl(d,doff)

    # Annotate each ACE with whether it grants dangerous write access
    for a in aces:
        m   =a["mask"]
        guid=a["object_guid"]

        # Always dangerous regardless of scope
        if m&GENERIC_ALL or m&WRITE_DACL or m&WRITE_OWNER or m&GENERIC_WRITE:
            full=(m&FULL_CTRL)==FULL_CTRL
            reason=("FullControl" if full else
                    "/".join(filter(None,[
                        "GenericAll"  if m&GENERIC_ALL   else "",
                        "WriteDACL"   if m&WRITE_DACL    else "",
                        "WriteOwner"  if m&WRITE_OWNER   else "",
                        "GenericWrite"if m&GENERIC_WRITE else "",
                    ])))
            a["dangerous_write"]=True; a["write_reason"]=reason
            continue

        if m&DS_WRITE_PROP:
            if not a["scoped"] or guid is None:
                # Unscoped WriteProperty (standard ACE, no GUID) = all attributes
                a["dangerous_write"]=True
                a["write_reason"]="WriteProperty (all attributes — unscoped)"
            elif guid in ENROLL_RIGHT_GUIDS:
                # Certificate-Enrollment / AutoEnrollment GUID with WriteProperty
                # is unusual but effectively just an enroll right — NOT a template write
                a["dangerous_write"]=False
                a["write_reason"]="WriteProperty[Enroll-right-GUID] (enroll only, not ESC4)"
            elif guid in DANGEROUS_ATTR_GUIDS:
                attr=GUID_NAMES.get(guid,guid)
                a["dangerous_write"]=True
                a["write_reason"]=f"WriteProperty[{attr}]"
            else:
                attr=GUID_NAMES.get(guid,guid)
                a["dangerous_write"]=False
                a["write_reason"]=f"WriteProperty[{attr}] (benign attribute)"
            continue

        a["dangerous_write"]=False; a["write_reason"]=""

    return aces

# ══════════════════════════════════════════════════════════
# MASK / RIGHTS
# ══════════════════════════════════════════════════════════

DS_READ_PROP=0x10; DS_WRITE_PROP=0x20; DS_CTL_ACCESS=0x100
DS_LIST_OBJ=0x80; DS_LIST=0x04
DELETE=0x10000; READ_CTRL=0x20000; WRITE_DACL=0x40000; WRITE_OWNER=0x80000
GENERIC_ALL=0x10000000; GENERIC_WRITE=0x40000000
AUTOENROLL=0x200
FULL_CTRL=(0x01|0x02|0x04|0x08|0x10|0x20|0x40|0x80|DELETE|READ_CTRL|WRITE_DACL|WRITE_OWNER)

def has_dangerous_write(ace):
    """True if this ACE grants write access that is exploitable for ESC4."""
    return ace.get("dangerous_write", False) and not ace["deny"]

def rights_labels(ace):
    """Human-readable rights for an ACE, using write_reason for accuracy."""
    m=ace["mask"]
    r=[]
    if ace.get("dangerous_write") and not ace["deny"]:
        wr=ace.get("write_reason","")
        if wr: r.append(wr)
    if m&DS_CTL_ACCESS:
        guid=ace.get("object_guid")
        if guid in ENROLL_RIGHT_GUIDS:
            lbl="AutoEnroll" if "Auto" in GUID_NAMES.get(guid,"") else "Enroll"
        elif guid:
            lbl=f"ExtRight[{GUID_NAMES.get(guid,guid[:8]+'…')}]"
        else:
            lbl="Enroll"
        r.append(lbl)
    if m&AUTOENROLL: r.append("AutoEnroll")
    if m&READ_CTRL:  r.append("ReadControl")
    if m&DS_READ_PROP: r.append("ReadProperty")
    if m&DS_LIST_OBJ:  r.append("ListObject")
    if m&DS_LIST:      r.append("ListContents")
    return r or ["(none)"]

# ══════════════════════════════════════════════════════════
# SID RESOLUTION
# ══════════════════════════════════════════════════════════

WKS={
    "S-1-1-0":("Everyone","low"),
    "S-1-5-7":("Anonymous Logon","low"),
    "S-1-5-11":("Authenticated Users","low"),
    "S-1-5-32-544":("BUILTIN\\Administrators","high"),
    "S-1-5-32-545":("BUILTIN\\Users","low"),
    "S-1-5-32-546":("BUILTIN\\Guests","low"),
    "S-1-5-18":("NT AUTHORITY\\SYSTEM","system"),
    "S-1-5-19":("NT AUTHORITY\\LOCAL SERVICE","system"),
    "S-1-5-20":("NT AUTHORITY\\NETWORK SERVICE","system"),
    "S-1-3-0":("Creator Owner","high"),
    "S-1-5-9":("Enterprise Domain Controllers","high"),
}
HIGH_RIDS={498,512,516,517,518,519,520,548,549,550,551}
LOW_RIDS={513,514,515,553}
_SC={}

def gval(e,a,d=None):
    try: v=e[a].value; return v if v is not None else d
    except: return d
def glist(e,a):
    try:
        v=e[a].value
        if v is None: return []
        return v if isinstance(v,list) else [v]
    except: return []

def resolve(conn, base, sid):
    if sid in _SC: return _SC[sid]
    if sid in WKS: _SC[sid]=WKS[sid]; return _SC[sid]
    try:
        conn.search(base,f"(objectSid={sid})",SUBTREE,
                    attributes=["sAMAccountName","objectClass","distinguishedName"])
        if conn.entries:
            e=conn.entries[0]
            sam=gval(e,"sAMAccountName",None)
            obj=glist(e,"objectClass")
            dn=(gval(e,"distinguishedName","") or "").lower()
            rid=int(sid.split("-")[-1]) if sid.count("-")>3 else None
            name=sam or sid
            if rid in HIGH_RIDS: tier="high"
            elif rid in LOW_RIDS: tier="low"
            elif "computer" in obj: tier="low"
            elif "group" in obj:
                hi=["domain admins","enterprise admins","schema admins",
                    "administrators","account operators","backup operators",
                    "read-only domain controllers","cert publishers"]
                lo=["domain users","domain computers","domain guests"]
                tier="high" if any(k in dn for k in hi) else \
                     "low"  if any(k in dn for k in lo) else "unknown"
            else: tier="unknown"
            _SC[sid]=(name,tier); return _SC[sid]
    except: pass
    rid=int(sid.split("-")[-1]) if sid.count("-")>3 else None
    tier="high" if rid in HIGH_RIDS else "low" if rid in LOW_RIDS else "unknown"
    _SC[sid]=(sid,tier); return _SC[sid]

# ══════════════════════════════════════════════════════════
# EKUs
# ══════════════════════════════════════════════════════════

EKU_MAP={
    "1.3.6.1.5.5.7.3.1":"Server Authentication",
    "1.3.6.1.5.5.7.3.2":"Client Authentication",
    "1.3.6.1.5.5.7.3.3":"Code Signing",
    "1.3.6.1.5.5.7.3.4":"Email Protection",
    "1.3.6.1.5.5.7.3.9":"OCSP Signing",
    "1.3.6.1.4.1.311.10.3.4":"EFS",
    "1.3.6.1.4.1.311.10.3.4.1":"EFS Recovery",
    "1.3.6.1.4.1.311.20.2.1":"Enrollment Agent",
    "1.3.6.1.4.1.311.20.2.2":"Smartcard Logon",
    "1.3.6.1.4.1.311.21.6":"Key Recovery Agent",
    "1.3.6.1.5.2.3.4":"PKINIT Client Auth",
    "1.3.6.1.5.2.3.5":"Kerberos Authentication",
    "2.5.29.37.0":"Any Purpose",
}
AUTH_EKUS={"1.3.6.1.5.5.7.3.2","1.3.6.1.4.1.311.20.2.2",
           "1.3.6.1.5.2.3.5","1.3.6.1.5.2.3.4","2.5.29.37.0"}

def eku_name(oid): return EKU_MAP.get(oid,oid)

# ══════════════════════════════════════════════════════════
# LDAP FETCHERS
# ══════════════════════════════════════════════════════════

T_ATTRS=["cn","displayName","msPKI-Certificate-Name-Flag","msPKI-Enrollment-Flag",
         "msPKI-RA-Signature","msPKI-Template-Schema-Version",
         "msPKI-Certificate-Application-Policy","pKIExtendedKeyUsage",
         "nTSecurityDescriptor","msPKI-Minimal-Key-Size","msPKI-Cert-Template-OID"]

def get_templates(conn,base):
    conn.search(f"CN=Certificate Templates,CN=Public Key Services,CN=Services,"
                f"CN=Configuration,{base}","(objectClass=pKICertificateTemplate)",
                SUBTREE,attributes=T_ATTRS,controls=SD_CONTROL)
    return conn.entries

def get_cas(conn,base):
    conn.search(f"CN=Enrollment Services,CN=Public Key Services,CN=Services,"
                f"CN=Configuration,{base}","(objectClass=pKIEnrollmentService)",
                SUBTREE,attributes=["cn","certificateTemplates","dNSHostName",
                "nTSecurityDescriptor","flags"],controls=SD_CONTROL)
    return conn.entries

def iflag(v):
    if v is None: return 0
    if isinstance(v,list): v=v[0]
    try: x=int(v); return x&0xFFFFFFFF if x<0 else x
    except: return 0

# ══════════════════════════════════════════════════════════
# ESC CHECKS
# ══════════════════════════════════════════════════════════

CT_SAN=0x1; CT_UPN=0x2; CT_PEND=0x2; CT_NOSEC=0x00080000

def run_checks(t, acl, ca_info=None, container_acl=None):
    """
    Returns ESC1-ESC9 check results.
    ca_info      : list of CA dicts from collect() — needed for ESC6/7/8
    container_acl: parsed ACL of the PKI container objects — needed for ESC5
    """
    nf=iflag(gval(t,"msPKI-Certificate-Name-Flag"))
    ef=iflag(gval(t,"msPKI-Enrollment-Flag"))
    eku=glist(t,"pKIExtendedKeyUsage")+glist(t,"msPKI-Certificate-Application-Policy")
    ra=iflag(gval(t,"msPKI-RA-Signature"))

    # Low/unknown-priv principals that can ENROLL (ESC1/2/3/9)
    lp=[(e["name"],e["rights"],e["tier"]) for e in acl
        if not e["deny"] and e["tier"] in("low","unknown")
        and({"Enroll","FullControl","WriteDACL","WriteOwner",
             "GenericWrite"}&set(e["rights"]))]

    # Non-admin principals with CONFIRMED dangerous write on template (ESC4)
    wp=[(e["name"],e["rights"],e["tier"],e.get("write_reason","")) for e in acl
        if not e["deny"] and e["tier"] not in("high","system")
        and e.get("dangerous_write",False)]

    # Non-admin principals with write on PKI containers (ESC5)
    cp=[]
    if container_acl:
        cp=[(e["name"],e["rights"],e["tier"]) for e in container_acl
            if not e["deny"] and e["tier"] not in("high","system")
            and({"FullControl","WriteDACL","WriteOwner","WriteProperty",
                 "GenericWrite"}&set(e["rights"]))]

    # CA-level flags (ESC6/7/8) — detected once and passed in
    c_esc6 = ca_info[0].get("esc6_flag",False) if ca_info else False
    c_esc7 = any(ca.get("esc7_vuln",False) for ca in (ca_info or []))
    c_esc8 = any(ca.get("esc8_accessible",False) for ca in (ca_info or []))

    c_san   =bool(nf&CT_SAN)
    c_upn   =bool(nf&CT_UPN)
    c_appr  =not bool(ef&CT_PEND)
    c_auth  =not eku or any(e in AUTH_EKUS for e in eku)
    c_any   =not eku or "2.5.29.37.0" in eku
    # ESC3: Enrollment Agent EKU is present explicitly,
    # OR Any Purpose is set (implies all EKUs incl. Enrollment Agent),
    # OR no EKUs defined (same as Any Purpose)
    c_ea    = ("1.3.6.1.4.1.311.20.2.1" in eku   # explicit Enrollment Agent OID
               or not eku                           # no EKU = any purpose = includes EA
               or "2.5.29.37.0" in eku)             # Any Purpose = includes EA
    c_nora  =ra==0
    c_lp    =len(lp)>0
    c_nosec =bool(ef&CT_NOSEC)

    def chk(esc, sev, conds, princs, notes=""):
        vuln=all(v for _,v in conds)
        return {"esc":esc,"sev":sev if vuln else "pass",
                "vuln":vuln,"conds":conds,
                "principals":princs if vuln else [],
                "notes": notes}

    return [
        chk("ESC1","critical",[
            ("Enrollee supplies SAN",c_san),
            ("No manager approval",c_appr),
            ("Auth EKU present",c_auth),
            ("Low/unknown priv can enroll",c_lp)],lp,
            "Direct domain compromise — enroll as Domain Admin immediately"),
        chk("ESC1-UPN","critical",[
            ("Enrollee supplies UPN SAN",c_upn),
            ("No manager approval",c_appr),
            ("Auth EKU present",c_auth),
            ("Low/unknown priv can enroll",c_lp)],lp,
            "UPN variant of ESC1 — same impact"),
        chk("ESC2","high",[
            ("Any Purpose / no EKU",c_any),
            ("No manager approval",c_appr),
            ("Low/unknown priv can enroll",c_lp)],lp,
            "Cert usable for any purpose — chain with ESC3 for domain compromise"),
        chk("ESC3","high",[
            ("Enrollment Agent EKU",c_ea),
            ("No manager approval",c_appr),
            ("No RA signature required",c_nora),
            ("Low/unknown priv can enroll",c_lp)],lp,
            "Get agent cert → enroll as DA on any other template"),
        chk("ESC4","high",[
            ("Non-admin has dangerous write rights (GUID-verified)",len(wp)>0)],
            [(n,r,t) for n,r,t,_ in wp],
            "Modify template flags → convert to ESC1 → enroll as DA"),
        chk("ESC5","high",[
            ("Non-admin has write on PKI container objects",len(cp)>0)],cp,
            "Control PKI containers → add/modify templates, CAs, or trust anchors"),
        chk("ESC6","high",[
            ("CA has EDITF_ATTRIBUTESUBJECTALTNAME2 flag",c_esc6)],[],
            "CA accepts SAN from CSR on ANY template — effectively ESC1 on all templates"),
        chk("ESC7","high",[
            ("Low-priv has ManageCA or ManageCertificates on CA",c_esc7)],[],
            "Issue certificates for any template or approve pending requests"),
        chk("ESC8","high",[
            ("Web enrollment accessible over HTTP (NTLM relay possible)",c_esc8)],[],
            "Relay any NTLM auth to CA web enrollment → get cert as any machine/user"),
        chk("ESC9","high",[
            ("NO_SECURITY_EXTENSION flag set",c_nosec),
            ("Auth EKU present",c_auth),
            ("Low/unknown priv can enroll",c_lp)],lp,
            "Missing SID binding — bypass strong certificate mapping"),
    ]

# ══════════════════════════════════════════════════════════
# RAW TOOL OUTPUT GENERATORS
# Per-ESC relevant tool output for every template.
# Covers ESC1-9 with appropriate commands and decoded output.
# ══════════════════════════════════════════════════════════

ACE_TYPE_NAMES = {
    0:"ACCESS_ALLOWED_ACE",1:"ACCESS_DENIED_ACE",
    5:"ACCESS_ALLOWED_OBJECT_ACE",6:"ACCESS_DENIED_OBJECT_ACE"
}

MASK_DACLEDIT = [
    (0x10000000,"GenericAll"),(0x40000000,"GenericWrite"),
    (0x00040000,"WriteDacl"),(0x00080000,"WriteOwner"),
    (0x00020000,"ReadControl"),(0x00010000,"Delete"),
    (0x00000100,"ControlAccess"),(0x00000080,"ListObject"),
    (0x00000040,"DeleteTree"),(0x00000020,"WriteProperty"),
    (0x00000010,"ReadProperty"),(0x00000008,"Self"),
    (0x00000004,"ListContents"),(0x00000002,"DeleteChild"),
    (0x00000001,"CreateChild"),
]

FULL_CTRL = (0x01|0x02|0x04|0x08|0x10|0x20|0x40|0x80|
             0x10000|0x20000|0x40000|0x80000)

def mask_to_dacledit_str(mask):
    if mask & 0x10000000: return "GenericAll"
    if (mask & FULL_CTRL) == FULL_CTRL: return "FullControl"
    parts = [n for bit,n in MASK_DACLEDIT if mask & bit]
    return ", ".join(parts) if parts else f"0x{mask:08x}"

def build_ldapsearch_output(tname, dn, sd_b64, raw_aces, conn, base):
    """Simulate ldapsearch output with decoded ACE breakdown."""
    lines = []
    lines.append(f"# Command:")
    ldap_cmd = (
        "ldapsearch -x -H ldap://DC_IP -D 'USER@DOMAIN' -w 'PASS' "
        f"-b '{dn}' '(objectClass=*)' nTSecurityDescriptor "
        "-E '!1.2.840.113556.1.4.801=::MAMCAQc='"
    )
    lines.append(ldap_cmd)
    lines.append("")
    lines.append(f"# Result:")
    lines.append(f"dn: {dn}")
    if sd_b64:
        # wrap at 70 chars like ldapsearch does
        b64_wrapped = "nTSecurityDescriptor:: "
        remaining = sd_b64
        first = True
        while remaining:
            if first:
                chunk = remaining[:47]
                remaining = remaining[47:]
                b64_wrapped += chunk
                first = False
            else:
                chunk = remaining[:70]
                remaining = remaining[70:]
                b64_wrapped += f"\n {chunk}"
        lines.append(b64_wrapped[:200] + " [... truncated for display]")
    lines.append("")
    lines.append("# Decoded ACEs:")
    for i, ace in enumerate(raw_aces):
        name_r, tier = resolve(conn, base, ace["sid"])
        atype_name = ACE_TYPE_NAMES.get(ace.get("atype",0), "ALLOW_OBJ" if ace.get("scoped") else "ALLOW")
        mask_str = mask_to_dacledit_str(ace["mask"])
        guid = ace.get("object_guid","")
        guid_name = GUID_NAMES.get(guid, "") if guid else ""
        lines.append(f"  ACE[{i}]: {atype_name}")
        lines.append(f"    Principal : {name_r} ({ace['sid']})")
        lines.append(f"    Mask      : 0x{ace['mask']:08x} = {mask_str}")
        if guid:
            lines.append(f"    GUID      : {guid}")
            lines.append(f"    GUID name : {guid_name or '(unknown)'}")
        if ace.get("dangerous_write") and not ace.get("deny"):
            lines.append(f"    *** ESC4 DANGEROUS: {ace.get('write_reason','')} ***")
        else:
            lines.append(f"    ESC4      : Not exploitable — {ace.get('write_reason','read/enroll only')}")
    return "\n".join(lines)

def build_dacledit_output(tname, dn, acl, conn, base):
    """Simulate impacket-dacledit -action read output."""
    lines = []
    lines.append(f"# Command:")
    lines.append(f"impacket-dacledit -action read -target-dn '{dn}' 'DOMAIN/USER:PASS' -dc-ip DC_IP")
    lines.append("")
    lines.append("# To filter by specific principal:")
    lines.append(f"impacket-dacledit -action read -principal 'Domain Users' -target-dn '{dn}' 'DOMAIN/USER:PASS' -dc-ip DC_IP")
    lines.append("")
    lines.append("Impacket v0.13.0 - Copyright Fortra, LLC")
    lines.append("[*] Parsing DACL")
    lines.append("[*] Printing parsed DACL")
    for i, e in enumerate(acl):
        if e["deny"]: continue  # show allows first
        mask_str = mask_to_dacledit_str(e["mask"])
        lines.append(f"[*]   ACE[{i}] info")
        lines.append(f"[*]     ACE Type                  : {'ACCESS_ALLOWED_OBJECT_ACE' if e.get('scoped') else 'ACCESS_ALLOWED_ACE'}")
        lines.append(f"[*]     ACE flags                 : None")
        lines.append(f"[*]     Access mask               : {mask_str} (0x{e['mask']:03x})")
        if e.get("scoped"):
            ace_key = next((a for a in [e] if a.get("sid")==e["sid"]), None)
            # find the original guid from raw aces via sid
            lines.append(f"[*]     Flags                     : ACE_OBJECT_TYPE_PRESENT")
            lines.append(f"[*]     Object type (GUID)        : {e.get('write_reason','Enrollment right')}")
        lines.append(f"[*]     Trustee (SID)             : {e['name']} ({e['sid']})")
        if e.get("dangerous_write"):
            lines.append(f"[*]     *** DANGEROUS WRITE: {e.get('write_reason','')} ***")
        else:
            lines.append(f"[*]     ESC4 verdict              : NOT exploitable")
    for i, e in enumerate(acl):
        if not e["deny"]: continue
        lines.append(f"[*]   ACE[{i}] DENY")
        lines.append(f"[*]     Trustee: {e['name']}  Rights: {mask_to_dacledit_str(e['mask'])}")
    lines.append("")
    lines.append("# ESC4 summary:")
    dangerous = [e for e in acl if e.get("dangerous_write") and not e["deny"]
                 and e["tier"] not in ("high","system")]
    if dangerous:
        for e in dangerous:
            lines.append(f"  *** [{e['tier'].upper()}] {e['name']} has dangerous write: {e.get('write_reason','')} ***")
    else:
        lines.append("  No non-admin principals have dangerous write access.")
        lines.append("  Any WriteProperty findings above are scoped to Enroll/AutoEnroll GUIDs (not ESC4).")
    return "\n".join(lines)

def build_bloodyad_output(tname, dn, raw_aces, conn, base):
    """Simulate bloodyAD get object --attr nTSecurityDescriptor output."""
    lines = []
    lines.append(f"# Command:")
    lines.append(f"bloodyAD -u USER -p 'PASS' -d DOMAIN --host DC_IP get object '{dn}' --attr nTSecurityDescriptor")
    lines.append("")
    lines.append(f"distinguishedName: {dn}")

    # Build SDDL string from raw aces
    sddl_parts = []
    for ace in raw_aces:
        name_r, tier = resolve(conn, base, ace["sid"])
        atype = "OA" if ace.get("scoped") else ("D" if ace.get("deny") else "A")
        mask = f"0x{ace['mask']:x}"
        guid = ace.get("object_guid","") or ""
        sid  = ace["sid"]
        sddl_parts.append(f"({atype};;{mask};{guid};;{sid})")
    sddl = f"O:{raw_aces[0]['sid'] if raw_aces else ''}G:...D:PAI" + "".join(sddl_parts) if raw_aces else "D:PAI"
    lines.append(f"nTSecurityDescriptor: {sddl}")
    lines.append("")
    lines.append("# SDDL breakdown:")
    lines.append("  O: = Owner SID")
    lines.append("  G: = Group SID")
    lines.append("  D: = DACL (access control list)")
    lines.append("  PAI = Protected + Auto-Inherited")
    lines.append("")
    lines.append("# Per-ACE breakdown:")
    for i, ace in enumerate(raw_aces):
        name_r, tier = resolve(conn, base, ace["sid"])
        atype_full = "ALLOW_OBJECT" if ace.get("scoped") else ("DENY" if ace.get("deny") else "ALLOW")
        mask_str = mask_to_dacledit_str(ace["mask"])
        guid = ace.get("object_guid","")
        guid_name = GUID_NAMES.get(guid,"") if guid else ""
        lines.append(f"  ACE {i+1}: ({('OA' if ace.get('scoped') else 'A')};;0x{ace['mask']:x};{guid};;{ace['sid']})")
        lines.append(f"    Type      : {atype_full}")
        lines.append(f"    Principal : {name_r} [{tier.upper()}]")
        lines.append(f"    Mask      : 0x{ace['mask']:08x} = {mask_str}")
        if guid:
            lines.append(f"    Scoped to : {guid} = {guid_name or '(unknown GUID)'}")
            lines.append(f"    Meaning   : {'Enroll/AutoEnroll right only — NOT WriteProperty on template attrs' if guid in ENROLL_RIGHT_GUIDS else ('DANGEROUS attr write' if guid in DANGEROUS_ATTR_GUIDS else 'Benign attribute')}")
        else:
            if ace.get("dangerous_write"):
                lines.append(f"    *** UNSCOPED: writes ALL attributes — ESC4 if non-admin ***")
            else:
                lines.append(f"    Scoped to : (none) — standard ACE")
        lines.append("")
    lines.append("# ESC4 verdict:")
    dangerous = [a for a in raw_aces if a.get("dangerous_write") and not a.get("deny")]
    if dangerous:
        for a in dangerous:
            name_r, tier = resolve(conn, base, a["sid"])
            if tier not in ("high","system"):
                lines.append(f"  *** REAL ESC4: [{tier.upper()}] {name_r} has {a.get('write_reason','')} ***")
        if not any(tier not in ("high","system")
                   for a in dangerous
                   for _, tier in [resolve(conn, base, a["sid"])]):
            lines.append("  Write access is only held by HIGH-tier admins (expected, not ESC4).")
    else:
        lines.append("  No dangerous write access found for any principal.")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════
# COLLECT ALL DATA
# ══════════════════════════════════════════════════════════

def fetch_container_acl(conn, base):
    """Fetch and parse the SD of the PKI containers for ESC5."""
    container_dns = [
        f"CN=Public Key Services,CN=Services,CN=Configuration,{base}",
        f"CN=Certificate Templates,CN=Public Key Services,CN=Services,CN=Configuration,{base}",
        f"CN=NTAuthCertificates,CN=Public Key Services,CN=Services,CN=Configuration,{base}",
        f"CN=AIA,CN=Public Key Services,CN=Services,CN=Configuration,{base}",
        f"CN=KRA,CN=Public Key Services,CN=Services,CN=Configuration,{base}",
    ]
    combined = []
    for dn in container_dns:
        try:
            conn.search(dn,"(objectClass=*)",attributes=["nTSecurityDescriptor"],
                        controls=SD_CONTROL)
            if conn.entries:
                sd = gval(conn.entries[0],"nTSecurityDescriptor")
                if sd:
                    aces = parse_sd(bytes(sd))
                    for ace in aces:
                        name_r, tier = resolve(conn, base, ace["sid"])
                        combined.append({
                            "container": dn.split(",")[0].replace("CN=",""),
                            "sid": ace["sid"], "name": name_r, "tier": tier,
                            "rights": rights_labels(ace), "deny": ace["deny"],
                            "dangerous_write": ace.get("dangerous_write",False),
                        })
        except Exception:
            pass
    return combined

def collect(conn, base):
    cas     = get_cas(conn, base)
    cmap    = {}
    ca_info = []

    # ESC7: check CA ACLs for low-priv write
    for ca in cas:
        cn=gval(ca,"cn","?"); dns=gval(ca,"dNSHostName","?")
        for t in glist(ca,"certificateTemplates"):
            cmap.setdefault(t,[]).append(cn)
        sd = gval(ca,"nTSecurityDescriptor")
        ca_acl = []
        if sd:
            for ace in parse_sd(bytes(sd)):
                name_r, tier = resolve(conn, base, ace["sid"])
                ca_acl.append({"name":name_r,"tier":tier,
                               "rights":rights_labels(ace),"deny":ace["deny"],
                               "dangerous_write":ace.get("dangerous_write",False)})
        esc7_vuln = any(e["dangerous_write"] and not e["deny"]
                        and e["tier"] not in ("high","system") for e in ca_acl)
        ca_info.append({
            "cn":cn,"dns":dns,"acl":ca_acl,
            "esc6_flag":False,    # requires registry read — marked manual
            "esc7_vuln":esc7_vuln,
            "esc8_accessible":False,  # requires HTTP check — marked manual
        })

    # ESC5: fetch PKI container ACLs
    container_acl = fetch_container_acl(conn, base)

    templates = get_templates(conn, base)
    results   = []

    for t in templates:
        name      = gval(t,"cn","?")
        disp      = gval(t,"displayName") or name
        schema    = gval(t,"msPKI-Template-Schema-Version","?")
        min_key   = gval(t,"msPKI-Minimal-Key-Size","?")
        nf        = iflag(gval(t,"msPKI-Certificate-Name-Flag"))
        ef        = iflag(gval(t,"msPKI-Enrollment-Flag"))
        ra        = iflag(gval(t,"msPKI-RA-Signature"))
        eku       = glist(t,"pKIExtendedKeyUsage")+glist(t,"msPKI-Certificate-Application-Policy")
        cas_pub   = cmap.get(name,[])
        published = len(cas_pub)>0
        approval  = bool(ef & CT_PEND)
        no_sec    = bool(ef & CT_NOSEC)
        san_flag  = bool(nf & CT_SAN)
        upn_flag  = bool(nf & CT_UPN)

        sd  = gval(t,"nTSecurityDescriptor")
        raw_aces = parse_sd(bytes(sd)) if sd else []

        # Merge ACEs per (sid, deny, object_guid) keeping dangerous_write from any ACE
        # Also group purely for display: same SID+deny → combine rights labels
        acl_merged = {}
        for ace in raw_aces:
            key = (ace["sid"], ace["deny"])
            if key not in acl_merged:
                acl_merged[key] = ace.copy()
                acl_merged[key]["_rights_list"] = rights_labels(ace)
            else:
                existing = acl_merged[key]
                # OR masks
                existing["mask"] |= ace["mask"]
                # propagate dangerous_write if any ACE for this SID is dangerous
                if ace.get("dangerous_write"):
                    existing["dangerous_write"] = True
                    existing["write_reason"] = ace.get("write_reason","")
                # merge labels (deduplicate)
                for lbl in rights_labels(ace):
                    if lbl not in existing["_rights_list"]:
                        existing["_rights_list"].append(lbl)
                # keep scoped=False if any unscoped ACE exists (more permissive)
                if not ace.get("scoped"):
                    existing["scoped"] = False

        acl = []
        for (sid, deny), ace in acl_merged.items():
            name_r, tier = resolve(conn, base, sid)
            rls = ace["_rights_list"]
            acl.append({
                "sid": sid, "name": name_r, "tier": tier,
                "rights": rls, "deny": deny,
                "mask": ace["mask"],
                "dangerous_write": ace.get("dangerous_write", False),
                "write_reason":    ace.get("write_reason", ""),
                "scoped":          ace.get("scoped", False),
            })

        checks = run_checks(t, acl, ca_info=ca_info, container_acl=container_acl)
        vulns  = [c for c in checks if c["vuln"]]
        max_sev= "pass"
        for c in checks:
            if c["sev"]=="critical": max_sev="critical"; break
            if c["sev"]=="high":     max_sev="high"
            if c["sev"]=="medium" and max_sev not in("high","critical"):
                max_sev="medium"

        # ── Generate raw tool output strings ──────────────────────────
        tmpl_dn = (f"CN={name},CN=Certificate Templates,CN=Public Key Services,"
                   f"CN=Services,CN=Configuration,{base}")

        # ldapsearch raw output simulation from actual SD bytes
        sd_b64 = ""
        if sd:
            import base64 as _b64
            sd_b64 = _b64.b64encode(bytes(sd)).decode()

        raw_ldapsearch   = build_ldapsearch_output(name, tmpl_dn, sd_b64, raw_aces, conn, base)
        raw_dacledit     = build_dacledit_output(name, tmpl_dn, acl, conn, base)
        raw_bloodyad     = build_bloodyad_output(name, tmpl_dn, raw_aces, conn, base)
        raw_tmpl_attrs   = build_template_attrs_output(
                               name, tmpl_dn, t, nf, ef, ra, eku)
        raw_ca_checks    = build_ca_checks_output(
                               name, ca_info, "DC_IP", "DOMAIN")

        results.append({
            "name":name,"display":disp,"schema":schema,"min_key":min_key,
            "published":published,"cas":cas_pub,"approval":approval,
            "no_sec":no_sec,"san_flag":san_flag,"upn_flag":upn_flag,
            "ra":ra,"eku":[eku_name(e) for e in eku],
            "acl":acl,"checks":checks,"vulns":vulns,"max_sev":max_sev,
            "raw_ldapsearch":  raw_ldapsearch,
            "raw_dacledit":    raw_dacledit,
            "raw_bloodyad":    raw_bloodyad,
            "raw_tmpl_attrs":  raw_tmpl_attrs,
            "raw_ca_checks":   raw_ca_checks,
            "tmpl_dn":         tmpl_dn,
        })

    return ca_info, results

# ══════════════════════════════════════════════════════════
# HTML GENERATION
# ══════════════════════════════════════════════════════════

def h(s): return html.escape(str(s))

def sev_cls(s):
    return {"critical":"sev-critical","high":"sev-high",
            "medium":"sev-medium","pass":"sev-pass"}.get(s,"sev-pass")

def tier_cls(t):
    return {"high":"tier-high","low":"tier-low",
            "system":"tier-system","unknown":"tier-unknown"}.get(t,"tier-unknown")

def build_html(ca_info, results, domain, dc_ip, diff_changes=None, prev_snap=None, curr_snap=None):
    diff_changes = diff_changes or []
    n_changes = len(diff_changes)
    n_critical = sum(1 for c in diff_changes if c.get("direction")=="WORSENED" or c["type"] in ("new_vuln","acl_added","rights_gained"))
    diff_badge = (f'<span class="diff-badge diff-badge-warn">{n_changes}</span>' if n_critical
                  else f'<span class="diff-badge">{n_changes}</span>' if n_changes
                  else '<span class="diff-badge diff-badge-ok">✓</span>')

    # ── Changes page content ──────────────────────────────────
    if prev_snap:
        changes_summary_html = (
            f"Comparing current scan ({curr_snap['ts'][:19] if curr_snap else 'now'}) "
            f"vs previous ({prev_snap['ts'][:19]}) · "
            f"<strong>{n_changes} change(s)</strong> detected"
        )
    else:
        changes_summary_html = "No previous scan found — run the tool again to see differences between scans."

    if not diff_changes:
        changes_content_html = '<div class="no-changes"><span class="no-changes-icon">✓</span><div>No changes detected vs previous scan. Environment is stable.</div></div>'
    else:
        TYPE_LABELS = {
            "new_template":    ("🆕", "New template",     "change-new"),
            "removed_template":("🗑", "Removed template", "change-removed"),
            "severity_change": ("⚡", "Severity change",  "change-sev"),
            "new_vuln":        ("🔴", "New vulnerability","change-vuln"),
            "fixed_vuln":      ("✅", "Fixed vulnerability","change-fixed"),
            "acl_added":       ("➕", "ACE added",         "change-acl"),
            "acl_removed":     ("➖", "ACE removed",       "change-acl"),
            "rights_gained":   ("⚠", "Rights gained",    "change-warn"),
            "rights_lost":     ("✓", "Rights removed",   "change-fixed"),
            "published_change":("📡", "Published change", "change-new"),
        }
        rows = ""
        for ch in diff_changes:
            icon, label, cls = TYPE_LABELS.get(ch["type"], ("?","Change","change-new"))
            tmpl  = h(ch.get("template",""))
            detail= h(ch.get("detail",""))
            princ = h(ch.get("principal",""))
            rights= h(", ".join(ch.get("rights",[])))
            direct= h(ch.get("direction",""))
            old_s = h(ch.get("old_sev","").upper())
            new_s = h(ch.get("new_sev","").upper())
            sev_arrow = f'<span class="sev-arrow">{old_s} → {new_s}</span>' if old_s and new_s else ""
            # Pre-compute conditionals (backslash not allowed in f-expr on Python <3.12)
            princ_div  = ('<div class="change-princ">Principal: ' + princ + '</div>') if princ else ''
            rights_div = ('<div class="change-rights">Rights: ' + rights + '</div>') if rights else ''
            rows += f"""
            <div class="change-row {cls}">
              <div class="change-icon">{icon}</div>
              <div class="change-body">
                <div class="change-label">{label}</div>
                <div class="change-template">{tmpl}</div>
                <div class="change-detail">{detail} {sev_arrow}</div>
                {princ_div}
                {rights_div}
              </div>
            </div>"""

        # Stats bar
        n_new_vuln  = sum(1 for c in diff_changes if c["type"]=="new_vuln")
        n_fixed     = sum(1 for c in diff_changes if c["type"]=="fixed_vuln")
        n_acl       = sum(1 for c in diff_changes if c["type"] in ("acl_added","rights_gained"))
        n_improved  = sum(1 for c in diff_changes if c.get("direction")=="IMPROVED")
        n_worsened  = sum(1 for c in diff_changes if c.get("direction")=="WORSENED")
        changes_content_html = f"""
        <div class="changes-stats">
          <div class="cs-card cs-red"><div class="cs-num">{n_new_vuln}</div><div class="cs-lbl">New Findings</div></div>
          <div class="cs-card cs-green"><div class="cs-num">{n_fixed}</div><div class="cs-lbl">Fixed</div></div>
          <div class="cs-card cs-orange"><div class="cs-num">{n_acl}</div><div class="cs-lbl">ACL Changes</div></div>
          <div class="cs-card cs-red"><div class="cs-num">{n_worsened}</div><div class="cs-lbl">Worsened</div></div>
          <div class="cs-card cs-green"><div class="cs-num">{n_improved}</div><div class="cs-lbl">Improved</div></div>
        </div>
        <div class="changes-list">{rows}</div>"""

    total   = len(results)
    pub     = sum(1 for r in results if r["published"])
    crits   = sum(1 for r in results if r["max_sev"]=="critical")
    highs   = sum(1 for r in results if r["max_sev"]=="high")
    meds    = sum(1 for r in results if r["max_sev"]=="medium")
    vulns_t = sum(1 for r in results if r["max_sev"]!="pass")

    # ── per-template cards
    cards = ""
    for r in sorted(results,
                    key=lambda x:({"critical":0,"high":1,"medium":2,"pass":3}
                                   .get(x["max_sev"],4), x["name"])):
        sev = r["max_sev"]

        # ACL rows
        acl_rows = ""
        for e in sorted(r["acl"], key=lambda x: x["tier"]):
            tc      = tier_cls(e["tier"])
            act     = "DENY" if e["deny"] else "ALLOW"
            act_cls = "deny" if e["deny"] else "allow"
            rstr    = ", ".join(e["rights"])
            is_dw   = e.get("dangerous_write",False) and not e["deny"]
            w_reason= h(e.get("write_reason",""))
            scope_tag = ""
            if e.get("scoped") and not e["deny"]:
                scope_tag = '<span class="scope-tag scoped-benign">scoped</span>'
                if is_dw:
                    scope_tag = '<span class="scope-tag scoped-danger">scoped-dangerous</span>'
            warn_html = f'<span class="warn-icon" title="{w_reason}">⚠ {w_reason}</span>' if is_dw else ""
            acl_rows += f"""
            <tr class="acl-row tier-row-{e['tier']}{'  dw-row' if is_dw else ''}">
              <td><span class="tier-badge {tc}">{h(e['tier'].upper())}</span></td>
              <td class="principal-cell">{h(e['name'])}</td>
              <td class="sid-cell">{h(e['sid'])}</td>
              <td><span class="action-badge {act_cls}">{act}</span></td>
              <td class="rights-cell">{h(rstr)} {scope_tag} {warn_html}</td>
            </tr>"""

        # ESC check rows
        esc_rows = ""
        for chk in r["checks"]:
            esc_sev = chk["sev"]
            verdict_html = f'<span class="verdict-badge {sev_cls(esc_sev)}">{h(esc_sev.upper())}</span>'
            cond_html = ""
            for cname, cval in chk["conds"]:
                icon = "✓" if cval else "✗"
                cls  = "cond-yes" if cval else "cond-no"
                cond_html += f'<div class="cond-row {cls}"><span class="cond-icon">{icon}</span>{h(cname)}</div>'
            princ_html = ""
            if chk["vuln"] and chk["principals"]:
                for pname, prights, ptier in chk["principals"][:5]:
                    if isinstance(prights, list): prights = ", ".join(prights)
                    ptc = tier_cls(ptier)
                    princ_html += f'<div class="princ-row"><span class="tier-badge {ptc} small">{h(ptier.upper())}</span> <strong>{h(str(pname))}</strong> → {h(str(prights))}</div>'
                if len(chk["principals"])>5:
                    princ_html += f'<div class="princ-more">…and {len(chk["principals"])-5} more</div>'

            esc_rows += f"""
            <tr class="esc-row">
              <td><span class="esc-label {sev_cls(esc_sev)}">{h(chk['esc'])}</span></td>
              <td>{cond_html}</td>
              <td>{verdict_html}</td>
              <td>{princ_html}</td>
            </tr>"""

        # EKU tags
        eku_tags = "".join(f'<span class="eku-tag">{h(e)}</span>' for e in r["eku"]) \
                   or '<span class="eku-tag eku-warn">None (Any Purpose)</span>'

        # CA tags
        ca_tags = "".join(f'<span class="ca-tag">{h(c)}</span>' for c in r["cas"]) \
                  or '<span class="unpub-tag">Not published</span>'

        # Flag pills
        flags = []
        if r["san_flag"]:  flags.append('<span class="flag-pill flag-warn">Enrollee Supplies SAN</span>')
        if r["upn_flag"]:  flags.append('<span class="flag-pill flag-warn">Enrollee Supplies UPN</span>')
        if r["approval"]:  flags.append('<span class="flag-pill flag-ok">Manager Approval Required</span>')
        if r["no_sec"]:    flags.append('<span class="flag-pill flag-warn">No Security Extension</span>')
        if not r["approval"]: flags.append('<span class="flag-pill flag-neutral">No Approval Needed</span>')
        flag_html = "".join(flags)

        # ── Verdict narrative
        verdict_blocks = ""
        for chk in r["checks"]:
            if not chk["vuln"]: continue
            esc    = chk["esc"]
            princs = chk.get("principals", [])
            pnames = [str(p[0]) for p in princs[:5]]
            extra  = len(princs) - 5 if len(princs) > 5 else 0
            plist  = ", ".join(f"<strong>{h(n)}</strong>" for n in pnames)
            if extra: plist += f" <em>and {extra} more</em>"
            tname  = h(r["name"])
            if esc in ("ESC1", "ESC1-UPN"):
                msg = (f"{plist} can request a certificate with an arbitrary Subject Alternative Name (SAN), "
                       f"allowing them to impersonate any domain user including Domain Admins. "
                       f"Immediately exploitable:<br>"
                       f"<code>certipy req -u USER -p PASS -ca CA -template {tname} -upn administrator@DOMAIN</code>")
            elif esc == "ESC2":
                msg = (f"{plist} can enroll a certificate with Any Purpose / no EKU, "
                       f"usable for client authentication, code signing, or any other purpose. "
                       f"Can be chained with other attacks to escalate privileges.")
            elif esc == "ESC3":
                msg = (f"{plist} can obtain an Enrollment Agent certificate and use it to "
                       f"request certificates on behalf of <em>any user</em>, including Domain Admins, "
                       f"against other templates that require agent-signed enrollment.")
            elif esc == "ESC4":
                msg = (f"{plist} {'has' if len(princs)==1 else 'have'} "
                       f"<em>write access</em> (FullControl / WriteDACL / WriteOwner / WriteProperty) "
                       f"over the <strong>{tname}</strong> template object. "
                       f"They can modify the template flags (e.g. enable <em>Enrollee Supplies Subject</em>) "
                       f"to convert this into an ESC1, then request a certificate as any user.<br>"
                       f"Exploit chain: <em>modify template</em> → <em>enroll with arbitrary SAN</em> → <em>authenticate as Domain Admin</em>")
            elif esc == "ESC9":
                msg = (f"{plist} can enroll a certificate that is missing the "
                       f"<em>szOID_NTDS_CA_SECURITY_EXT</em> SID binding. "
                       f"Combined with shadow credentials or NTLM relay, "
                       f"this allows authentication as another principal without the certificate being tied to the requestor SID.")
            else:
                msg = f"{plist} meet all conditions for {h(esc)}."
            sev_c = sev_cls(chk["sev"])
            verdict_blocks += (
                f'''<div class="verdict-block {sev_c}-block">'''
                f'''<div class="verdict-title">'''
                f'''<span class="verdict-esc-label {sev_c}">{h(esc)}</span>'''
                f'''<span class="verdict-headline">Exploitation path confirmed — all conditions met</span>'''
                f'''</div><div class="verdict-text">{msg}</div></div>'''
            )

        # Vuln count badge
        vuln_count = len(r["vulns"])
        count_badge = f'<span class="vuln-count {sev_cls(sev)}">{vuln_count} finding{"s" if vuln_count!=1 else ""}</span>' \
                      if vuln_count else '<span class="vuln-count sev-pass">Clean</span>'

        # ── Principal permissions analysis ─────────────────────────
        # Group all ACE entries by principal, show what they have
        # and whether it meets ESC requirements
        princ_map = {}
        for e in r["acl"]:
            k = (e["name"], e["tier"], e["sid"])
            if k not in princ_map:
                princ_map[k] = {"allow": [], "deny": [], "dangerous": False, "write_reason": ""}
            rls = e["rights"]
            if e["deny"]:
                princ_map[k]["deny"].extend(rls)
            else:
                princ_map[k]["allow"].extend(rls)
                if e.get("dangerous_write"):
                    princ_map[k]["dangerous"] = True
                    princ_map[k]["write_reason"] = e.get("write_reason","")

        # ESC conditions already stored in r
        # r["san_flag"], r["approval"], r["checks"] etc.
        NEEDED_ESC1  = ["Enroll","FullControl","WriteDACL","WriteOwner","GenericWrite"]
        NEEDED_ESC4  = ["FullControl","WriteDACL","WriteOwner","GenericWrite",
                        "WriteProperty (all attributes — unscoped)",
                        "WriteProperty[msPKI-Certificate-Name-Flag]",
                        "WriteProperty[msPKI-Enrollment-Flag]",
                        "WriteProperty[pKIExtendedKeyUsage]",
                        "WriteProperty[msPKI-Certificate-Application-Policy]",
                        "WriteProperty[msPKI-RA-Signature]"]

        pa_rows = ""
        for (pname, ptier, psid), pdata in sorted(
                princ_map.items(), key=lambda x: x[0][1]):  # sort by tier
            allow_set = set(pdata["allow"])
            tc  = tier_cls(ptier)

            # What they CAN do
            can_enroll  = bool({"Enroll","AutoEnroll","FullControl"} & allow_set)
            can_write   = pdata["dangerous"]
            write_r     = pdata["write_reason"]

            # Missing for ESC1: need Enroll + (other template conditions)
            # Missing for ESC4: need dangerous write
            missing_esc4 = []
            if not can_write:
                missing_esc4.append("No dangerous WriteProperty/WriteDACL/WriteOwner/GenericAll")

            have_html = "".join(
                f'<span class="perm-pill perm-{"write" if r in NEEDED_ESC4 else "enroll" if "Enroll" in r or "AutoEnroll" in r else "read"}">{h(r)}</span>'
                for r in sorted(allow_set) if r not in ["(none)"]
            ) or '<span class="perm-pill perm-read">No rights</span>'

            deny_html = ""
            if pdata["deny"]:
                deny_html = ' '.join(f'<span class="perm-pill perm-deny">{h(r)}</span>'
                                     for r in pdata["deny"])

            # ESC4 verdict for this principal
            esc4_icon = "✓" if can_write else "✗"
            esc4_cls  = "pa-yes" if can_write else "pa-no"
            esc1_icon = "✓" if can_enroll else "✗"
            esc1_cls  = "pa-yes" if can_enroll else "pa-no"

            missing_html = ""
            if not can_write:
                missing_html = '<div class="pa-missing">Missing for ESC4: dangerous write (WriteDACL / WriteOwner / GenericAll / WriteProperty on msPKI attrs)</div>'

            pa_rows += f"""
            <div class="pa-row tier-bg-{ptier}">
              <div class="pa-header">
                <span class="tier-badge {tc}">{h(ptier.upper())}</span>
                <span class="pa-name">{h(pname)}</span>
                <span class="pa-sid">{h(psid)}</span>
                <div class="pa-verdict-pills">
                  <span class="pa-verdict {esc1_cls}" title="Can enroll on this template">
                    {esc1_icon} Enroll
                  </span>
                  <span class="pa-verdict {esc4_cls}" title="Has dangerous write for ESC4">
                    {esc4_icon} ESC4 write
                  </span>
                </div>
              </div>
              <div class="pa-rights">
                <span class="pa-label">Allowed:</span>{have_html}
                {'<span class="pa-label">Denied:</span>'+deny_html if deny_html else ""}
              </div>
              {missing_html}
            </div>"""

        princ_analysis_html = pa_rows if pa_rows else '<div class="pa-empty">No ACEs parsed</div>'

        verdict_section = (
            '<div class="verdict-section">' + verdict_blocks + '</div>'
        ) if verdict_blocks else ""

        # Pre-compute tiers for data attribute (set literal not allowed in f-string)
        _write_rights = {"Enroll","FullControl","WriteDACL","WriteOwner","WriteProperty","GenericWrite","AutoEnroll"}
        _tiers_str = ",".join(sorted(set(
            e["tier"] for e in r["acl"]
            if not e["deny"] and _write_rights & set(e["rights"])
        )))

        cards += f"""
        <div class="template-card" data-sev="{sev}" data-published="{str(r['published']).lower()}"
             data-name="{h(r['name']).lower()}"
             data-escs="{','.join(c['esc'].lower().replace(' ','').replace('-upn','1-upn') for c in r['checks'] if c['vuln'])}"
             data-tiers="{_tiers_str}">
          <div class="card-header sev-border-{sev}" onclick="toggleCard(this)">
            <div class="card-title">
              <span class="sev-dot {sev_cls(sev)}"></span>
              <span class="tmpl-name">{h(r['name'])}</span>
              {count_badge}
            </div>
            <div class="card-meta">
              <span class="meta-item">📋 Schema v{h(r['schema'])}</span>
              <span class="meta-item">🔑 {h(r['min_key'])} bits</span>
              <span class="meta-item">📡 {'Published' if r['published'] else 'Not published'}</span>
              <span class="chevron">›</span>
            </div>
          </div>
          <div class="card-body" style="display:none">

            <div class="info-grid">
              <div class="info-block">
                <div class="info-label">Published On</div>
                <div>{ca_tags}</div>
              </div>
              <div class="info-block">
                <div class="info-label">EKUs</div>
                <div>{eku_tags}</div>
              </div>
              <div class="info-block">
                <div class="info-label">Flags</div>
                <div>{flag_html}</div>
              </div>
              <div class="info-block">
                <div class="info-label">RA Signatures Required</div>
                <div class="info-val">{h(r['ra'])}</div>
              </div>
            </div>

            <div class="section-title">ACL — All Permissions</div>
            <div class="table-wrap">
              <table class="acl-table">
                <thead><tr>
                  <th>Tier</th><th>Principal</th><th>SID</th>
                  <th>Action</th><th>Rights</th>
                </tr></thead>
                <tbody>{acl_rows}</tbody>
              </table>
            </div>

            <div class="section-title">Principal Permissions Analysis</div>
            <div class="princ-analysis">
              {princ_analysis_html}
            </div>

            <div class="section-title">ESC Vulnerability Checks</div>
            <div class="table-wrap">
              <table class="esc-table">
                <thead><tr>
                  <th>ESC</th><th>Conditions</th><th>Verdict</th><th>Triggering Principals</th>
                </tr></thead>
                <tbody>{esc_rows}</tbody>
              </table>
            </div>

          {verdict_section}

          <div class="raw-tools-section">
            <div class="raw-tools-header">
              <span class="raw-tools-title">🔬 Raw Tool Verification</span>
              <span class="raw-tools-sub">Commands and decoded output for every ESC type — run independently to verify</span>
            </div>
            <div class="raw-tools-tabs">
              <button class="raw-tab active" onclick="switchRawTab(this, '{h(r['name'])}-attrs')">ESC1/2/3/9 — Template Flags</button>
              <button class="raw-tab" onclick="switchRawTab(this, '{h(r['name'])}-ldap')">ESC4 — ldapsearch ACL</button>
              <button class="raw-tab" onclick="switchRawTab(this, '{h(r['name'])}-dacl')">ESC4 — dacledit</button>
              <button class="raw-tab" onclick="switchRawTab(this, '{h(r['name'])}-bloody')">ESC4 — bloodyAD</button>
              <button class="raw-tab" onclick="switchRawTab(this, '{h(r['name'])}-ca')">ESC5/6/7/8 — CA Checks</button>
            </div>

            <div id="{h(r['name'])}-attrs" class="raw-panel active-panel">
              <div class="raw-explain">
                <strong>ESC1/2/3/9 — Template Attributes</strong> via ldapsearch.
                Shows decoded <code>msPKI-Certificate-Name-Flag</code>, <code>msPKI-Enrollment-Flag</code>,
                <code>pKIExtendedKeyUsage</code> and <code>msPKI-RA-Signature</code>
                with per-bit analysis showing exactly which ESC conditions each flag satisfies.
              </div>
              <pre class="raw-output">{h(r['raw_tmpl_attrs'])}</pre>
            </div>

            <div id="{h(r['name'])}-ldap" class="raw-panel" style="display:none">
              <div class="raw-explain">
                <strong>ESC4 — ldapsearch</strong> fetches the raw <code>nTSecurityDescriptor</code>.
                Decoded with every ACE showing its ObjectType GUID.
                If GUID = <code>0e10c968-...</code> (Certificate-Enrollment) → <em>Enroll right only, NOT ESC4</em>.
                Dangerous: unscoped <code>WriteProperty</code>, <code>WriteDACL</code>, <code>WriteOwner</code>, <code>GenericAll</code>.
              </div>
              <pre class="raw-output">{h(r['raw_ldapsearch'])}</pre>
            </div>

            <div id="{h(r['name'])}-dacl" class="raw-panel" style="display:none">
              <div class="raw-explain">
                <strong>ESC4 — impacket-dacledit</strong> output format.
                <code>ACCESS_ALLOWED_OBJECT_ACE</code> with GUID = scoped right (check GUID meaning).
                <code>ACCESS_ALLOWED_ACE</code> without GUID = unscoped = dangerous if mask includes WriteProperty.
                Empty output for a principal = <em>no direct ACE</em> (access only through group membership).
              </div>
              <pre class="raw-output">{h(r['raw_dacledit'])}</pre>
            </div>

            <div id="{h(r['name'])}-bloody" class="raw-panel" style="display:none">
              <div class="raw-explain">
                <strong>ESC4 — bloodyAD</strong> SDDL format: <code>(TYPE;;MASK;GUID;;SID)</code>.
                <code>OA</code> = Object ACE (scoped). <code>A</code> = Standard ACE (unscoped, more dangerous).
                Mask <code>0x130</code> + GUID <code>0e10c968-...</code> = Enroll right only.
                Mask <code>0xf00ff</code> no GUID = FullControl.
              </div>
              <pre class="raw-output">{h(r['raw_bloodyad'])}</pre>
            </div>

            <div id="{h(r['name'])}-ca" class="raw-panel" style="display:none">
              <div class="raw-explain">
                <strong>ESC5/6/7/8 — CA-level checks</strong>.
                ESC5: write on PKI containers. ESC6: CA registry flag (certutil).
                ESC7: ManageCA/ManageCertificates rights. ESC8: HTTP web enrollment endpoint.
                Copy and run the commands below to verify each finding independently.
              </div>
              <pre class="raw-output">{h(r['raw_ca_checks'])}</pre>
            </div>
          </div>

          </div>
        </div>"""

    # ── CA info
    ca_rows = ""
    for ca in ca_info:
        ca_rows += f"""<tr>
          <td>{h(ca['cn'])}</td>
          <td>{h(ca['dns'])}</td>
          <td><a class="ext-link" href="http://{h(ca['dns'])}/certsrv/certfnsh.asp" target="_blank">Check ESC8</a></td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>ADCS Audit — {h(domain)}</title>
<style>
:root {{
  --bg:       #0d1117;
  --bg2:      #161b22;
  --bg3:      #21262d;
  --border:   #30363d;
  --text:     #e6edf3;
  --text2:    #8b949e;
  --red:      #f85149;
  --orange:   #e3b341;
  --yellow:   #d29922;
  --green:    #3fb950;
  --blue:     #58a6ff;
  --purple:   #bc8cff;
  --cyan:     #39d353;
  --font:     'JetBrains Mono', 'Fira Mono', 'Cascadia Code', monospace;
}}
* {{ box-sizing:border-box; margin:0; padding:0 }}
body {{ background:var(--bg); color:var(--text); font-family:var(--font);
        font-size:13px; line-height:1.6 }}
a {{ color:var(--blue); text-decoration:none }}
a:hover {{ text-decoration:underline }}

/* ── Layout */
.topbar {{ background:var(--bg2); border-bottom:1px solid var(--border);
           padding:10px 28px; display:flex; align-items:center;
           justify-content:space-between; position:sticky; top:0; z-index:100 }}
.topbar-left {{ display:flex; flex-direction:column; gap:1px }}
.topbar-title {{ font-size:15px; font-weight:700; color:var(--text); letter-spacing:.5px }}
.topbar-nav {{ display:flex; gap:6px }}
.nav-btn {{ background:transparent; border:1px solid var(--border); color:var(--text2);
            padding:6px 18px; border-radius:6px; cursor:pointer;
            font-family:var(--font); font-size:12px; transition:all .15s }}
.nav-btn:hover {{ border-color:var(--blue); color:var(--text) }}
.nav-btn.active {{ border-color:var(--blue); color:var(--blue);
                   background:rgba(88,166,255,.12); font-weight:700 }}
.domain {{ color:var(--blue); font-size:12px }}
.export-btns {{ display:flex; gap:6px; align-items:center }}
.export-btn {{ background:rgba(88,166,255,.1); border:1px solid rgba(88,166,255,.3);
               color:var(--blue); padding:5px 14px; border-radius:6px;
               font-family:var(--font); font-size:11px; font-weight:700;
               text-decoration:none; transition:all .15s; cursor:pointer }}
.export-btn:hover {{ background:rgba(88,166,255,.2); border-color:var(--blue);
                     text-decoration:none }}
.export-btn-sec {{ background:rgba(139,148,158,.08);
                   border-color:rgba(139,148,158,.3); color:var(--text2) }}
.export-btn-sec:hover {{ color:var(--text); border-color:var(--text2) }}
.diff-badge {{ font-size:10px; font-weight:700; padding:1px 6px;
               border-radius:8px; margin-left:4px;
               background:rgba(139,148,158,.2); color:var(--text2) }}
.diff-badge-warn {{ background:rgba(248,81,73,.2); color:var(--red) }}
.diff-badge-ok   {{ background:rgba(63,185,80,.2);  color:var(--green) }}
/* ── Changes page ── */
.no-changes {{ display:flex; align-items:center; gap:16px; background:rgba(63,185,80,.08);
               border:1px solid rgba(63,185,80,.2); border-radius:8px;
               padding:24px; font-size:14px; color:var(--green); margin:20px 0 }}
.no-changes-icon {{ font-size:32px }}
.changes-stats {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(120px,1fr));
                  gap:10px; margin-bottom:20px }}
.cs-card {{ background:var(--bg2); border:1px solid var(--border); border-radius:8px;
            padding:14px; text-align:center }}
.cs-num {{ font-size:28px; font-weight:700 }}
.cs-lbl {{ font-size:10px; text-transform:uppercase; letter-spacing:.5px;
           color:var(--text2); margin-top:3px }}
.cs-red .cs-num {{ color:var(--red) }}
.cs-green .cs-num {{ color:var(--green) }}
.cs-orange .cs-num {{ color:var(--orange) }}
.changes-list {{ display:flex; flex-direction:column; gap:8px }}
.change-row {{ display:flex; gap:14px; background:var(--bg2);
               border:1px solid var(--border); border-radius:8px;
               padding:14px; align-items:flex-start }}
.change-new     {{ border-left:3px solid var(--blue) }}
.change-removed {{ border-left:3px solid var(--text2) }}
.change-sev     {{ border-left:3px solid var(--orange) }}
.change-vuln    {{ border-left:3px solid var(--red) }}
.change-fixed   {{ border-left:3px solid var(--green) }}
.change-acl     {{ border-left:3px solid var(--yellow) }}
.change-warn    {{ border-left:3px solid var(--red) }}
.change-icon {{ font-size:20px; flex-shrink:0; width:28px; text-align:center }}
.change-body {{ flex:1 }}
.change-label  {{ font-size:10px; font-weight:700; text-transform:uppercase;
                  letter-spacing:.6px; color:var(--text2); margin-bottom:3px }}
.change-template {{ font-size:14px; font-weight:700; margin-bottom:4px }}
.change-detail {{ font-size:12px; color:var(--text2) }}
.change-princ  {{ font-size:11px; color:var(--text2); margin-top:4px }}
.change-rights {{ font-size:11px; color:var(--orange); margin-top:2px }}
.sev-arrow {{ background:var(--bg3); padding:2px 8px; border-radius:4px;
              font-size:11px; font-weight:700; margin-left:8px }}
.main {{ max-width:1400px; margin:0 auto; padding:24px 28px }}

/* ── Stats */
.stats {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr));
          gap:12px; margin-bottom:28px }}
.stat-card {{ background:var(--bg2); border:1px solid var(--border);
              border-radius:8px; padding:16px; text-align:center }}
.stat-num {{ font-size:32px; font-weight:700; line-height:1 }}
.stat-label {{ font-size:11px; color:var(--text2); margin-top:4px; text-transform:uppercase; letter-spacing:.5px }}
.num-critical {{ color:var(--red) }}
.num-high     {{ color:var(--orange) }}
.num-medium   {{ color:var(--yellow) }}
.num-pass     {{ color:var(--green) }}
.num-total    {{ color:var(--blue) }}

/* ── Filter bar */
.filter-bar {{ display:flex; gap:8px; margin-bottom:20px; flex-wrap:wrap; align-items:flex-start }}
.filter-bar input {{ background:var(--bg2); border:1px solid var(--border);
                     color:var(--text); padding:7px 12px; border-radius:6px;
                     font-family:var(--font); font-size:12px; width:220px; align-self:center }}
.filter-bar input:focus {{ outline:none; border-color:var(--blue) }}
.filter-group {{ display:flex; flex-wrap:wrap; align-items:center; gap:5px;
                 background:var(--bg2); border:1px solid var(--border);
                 border-radius:8px; padding:6px 10px }}
.filter-group-label {{ font-size:9px; font-weight:700; letter-spacing:.8px;
                        color:var(--text2); text-transform:uppercase;
                        margin-right:4px; align-self:center }}
.filter-btn {{ background:transparent; border:1px solid var(--border);
               color:var(--text2); padding:4px 11px; border-radius:5px;
               cursor:pointer; font-family:var(--font); font-size:11px;
               transition:all .15s }}
.filter-btn:hover {{ border-color:var(--blue); color:var(--text) }}
.sev-btn.active  {{ border-color:var(--blue);   color:var(--blue);   background:rgba(88,166,255,.12) }}
.esc-btn.active  {{ border-color:var(--orange); color:var(--orange); background:rgba(227,179,65,.12) }}
.tier-btn.active {{ border-color:var(--green);  color:var(--green);  background:rgba(63,185,80,.12) }}
.count-label {{ color:var(--text2); font-size:12px; align-self:center; margin-left:4px }}

/* ── Template cards */
.template-card {{ background:var(--bg2); border:1px solid var(--border);
                  border-radius:8px; margin-bottom:10px; overflow:hidden }}
.card-header {{ display:flex; justify-content:space-between; align-items:center;
                padding:13px 18px; cursor:pointer; user-select:none }}
.card-header:hover {{ background:var(--bg3) }}
.card-title {{ display:flex; align-items:center; gap:10px }}
.sev-dot {{ width:8px; height:8px; border-radius:50%; flex-shrink:0 }}
.sev-dot.sev-critical {{ background:var(--red) }}
.sev-dot.sev-high     {{ background:var(--orange) }}
.sev-dot.sev-medium   {{ background:var(--yellow) }}
.sev-dot.sev-pass     {{ background:var(--green) }}
.tmpl-name {{ font-weight:700; font-size:13px }}
.card-meta {{ display:flex; align-items:center; gap:14px; color:var(--text2); font-size:11px }}
.meta-item {{ display:flex; align-items:center; gap:4px }}
.chevron {{ font-size:18px; transition:transform .2s; color:var(--text2) }}
.chevron.open {{ transform:rotate(90deg) }}
.sev-border-critical {{ border-left:3px solid var(--red) }}
.sev-border-high     {{ border-left:3px solid var(--orange) }}
.sev-border-medium   {{ border-left:3px solid var(--yellow) }}
.sev-border-pass     {{ border-left:3px solid var(--green) }}
.card-body {{ padding:0 18px 18px }}

/* ── Info grid */
.info-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(260px,1fr));
              gap:14px; margin:14px 0 }}
.info-block {{ background:var(--bg3); border:1px solid var(--border);
               border-radius:6px; padding:12px }}
.info-label {{ font-size:10px; text-transform:uppercase; letter-spacing:.6px;
               color:var(--text2); margin-bottom:6px }}
.info-val {{ font-size:13px }}

/* ── Tags / badges */
.tier-badge {{ font-size:10px; font-weight:700; padding:2px 7px;
               border-radius:4px; letter-spacing:.5px }}
.tier-badge.small {{ font-size:9px; padding:1px 5px }}
.tier-high    {{ background:rgba(227,179,65,.15);  color:var(--orange) }}
.tier-low     {{ background:rgba(248,81,73,.15);   color:var(--red) }}
.tier-system  {{ background:rgba(88,166,255,.15);  color:var(--blue) }}
.tier-unknown {{ background:rgba(139,148,158,.15); color:var(--text2) }}
.action-badge {{ font-size:10px; font-weight:700; padding:2px 7px; border-radius:4px }}
.allow {{ background:rgba(63,185,80,.15); color:var(--green) }}
.deny  {{ background:rgba(248,81,73,.15); color:var(--red) }}
.eku-tag {{ background:rgba(88,166,255,.1); color:var(--blue); font-size:11px;
            padding:2px 8px; border-radius:4px; margin:2px; display:inline-block }}
.eku-warn {{ background:rgba(248,81,73,.1); color:var(--red) }}
.ca-tag {{ background:rgba(63,185,80,.1); color:var(--green); font-size:11px;
           padding:2px 8px; border-radius:4px; margin:2px; display:inline-block }}
.unpub-tag {{ background:rgba(248,81,73,.1); color:var(--red); font-size:11px;
              padding:2px 8px; border-radius:4px }}
.flag-pill {{ font-size:10px; padding:2px 8px; border-radius:12px;
              margin:2px; display:inline-block; font-weight:600 }}
.flag-warn    {{ background:rgba(248,81,73,.15); color:var(--red) }}
.flag-ok      {{ background:rgba(63,185,80,.15); color:var(--green) }}
.flag-neutral {{ background:rgba(139,148,158,.1); color:var(--text2) }}
.vuln-count {{ font-size:10px; font-weight:700; padding:2px 9px;
               border-radius:10px; border:1px solid }}
.sev-critical {{ color:var(--red);    border-color:var(--red);    background:rgba(248,81,73,.1) }}
.sev-high     {{ color:var(--orange); border-color:var(--orange); background:rgba(227,179,65,.1) }}
.sev-medium   {{ color:var(--yellow); border-color:var(--yellow); background:rgba(210,153,34,.1) }}
.sev-pass     {{ color:var(--green);  border-color:var(--green);  background:rgba(63,185,80,.1) }}
.verdict-badge {{ font-size:11px; font-weight:700; padding:3px 10px; border-radius:4px }}
.esc-label {{ font-size:11px; font-weight:700; padding:3px 8px; border-radius:4px;
              white-space:nowrap }}
.warn-icon {{ color:var(--orange) }}

/* ── Tables */
.section-title {{ font-size:11px; font-weight:700; text-transform:uppercase;
                  letter-spacing:.8px; color:var(--text2); margin:18px 0 8px;
                  border-bottom:1px solid var(--border); padding-bottom:6px }}
.table-wrap {{ overflow-x:auto; border-radius:6px; border:1px solid var(--border) }}
table {{ width:100%; border-collapse:collapse }}
th {{ background:var(--bg3); color:var(--text2); font-size:10px; font-weight:700;
      text-transform:uppercase; letter-spacing:.5px; padding:8px 12px;
      text-align:left; border-bottom:1px solid var(--border) }}
td {{ padding:7px 12px; border-bottom:1px solid rgba(48,54,61,.5);
      vertical-align:top; font-size:12px }}
tr:last-child td {{ border-bottom:none }}
.acl-table tr:hover {{ background:var(--bg3) }}
.principal-cell {{ font-weight:600 }}
.sid-cell {{ color:var(--text2); font-size:11px }}
.rights-cell {{ color:var(--text); max-width:320px }}
.cond-row {{ display:flex; align-items:center; gap:6px;
             padding:2px 0; font-size:12px }}
.cond-yes {{ color:var(--green) }}
.cond-no  {{ color:var(--text2) }}
.cond-icon {{ font-size:12px; width:14px }}
.princ-row {{ font-size:11px; padding:2px 0; display:flex; align-items:center; gap:6px }}
.princ-more {{ font-size:11px; color:var(--text2); font-style:italic; padding-top:2px }}
.esc-table td {{ padding:10px 12px }}

/* ── CA table */
.ca-section {{ background:var(--bg2); border:1px solid var(--border);
               border-radius:8px; margin-bottom:24px; overflow:hidden }}
.ca-section h2 {{ padding:12px 18px; font-size:12px; font-weight:700;
                  text-transform:uppercase; letter-spacing:.8px;
                  background:var(--bg3); border-bottom:1px solid var(--border) }}
.ext-link {{ color:var(--blue); font-size:11px }}

/* ── Legend */
.legend {{ display:flex; gap:20px; flex-wrap:wrap; padding:16px 0; color:var(--text2);
           font-size:11px; border-top:1px solid var(--border); margin-top:28px }}
.legend-item {{ display:flex; align-items:center; gap:6px }}
.filter-divider {{ width:1px; height:28px; background:var(--border); margin:0 4px; flex-shrink:0 }}
.verdict-section {{ margin-top:18px; padding-top:4px }}
.verdict-block {{ border-radius:6px; padding:14px 16px; margin-bottom:10px; border-left:3px solid }}
.sev-critical-block {{ background:rgba(248,81,73,.07); border-color:var(--red) }}
.sev-high-block     {{ background:rgba(227,179,65,.07); border-color:var(--orange) }}
.sev-medium-block   {{ background:rgba(210,153,34,.07); border-color:var(--yellow) }}
.verdict-title {{ display:flex; align-items:center; gap:10px; margin-bottom:8px }}
.verdict-esc-label {{ font-size:11px; font-weight:700; padding:2px 8px; border-radius:4px; flex-shrink:0 }}
.verdict-headline {{ font-size:12px; font-weight:700; color:var(--text) }}
.verdict-text {{ font-size:12px; color:var(--text2); line-height:1.8 }}
.verdict-text strong {{ color:var(--text) }}
.verdict-text code {{ background:var(--bg3); color:var(--cyan); padding:2px 7px; border-radius:4px; font-size:11px }}
.verdict-text em {{ color:var(--orange); font-style:normal; font-weight:600 }}
.scope-tag {{ font-size:9px; padding:1px 6px; border-radius:3px; margin-left:4px; font-weight:700; letter-spacing:.4px }}
.scoped-benign {{ background:rgba(139,148,158,.15); color:var(--text2) }}
.scoped-danger {{ background:rgba(248,81,73,.15); color:var(--red) }}
.dw-row {{ background:rgba(248,81,73,.04) }}
.warn-icon {{ color:var(--orange); font-size:11px }}
/* ── Principal analysis ── */
.princ-analysis {{ border:1px solid var(--border); border-radius:6px;
                   overflow:hidden; margin-bottom:4px }}
.pa-row {{ padding:10px 14px; border-bottom:1px solid rgba(48,54,61,.4) }}
.pa-row:last-child {{ border-bottom:none }}
.tier-bg-low     {{ background:rgba(248,81,73,.03) }}
.tier-bg-unknown {{ background:rgba(139,148,158,.03) }}
.tier-bg-high    {{ background:rgba(227,179,65,.03) }}
.tier-bg-system  {{ background:rgba(88,166,255,.03) }}
.pa-header {{ display:flex; align-items:center; gap:8px; flex-wrap:wrap; margin-bottom:6px }}
.pa-name {{ font-weight:700; font-size:12px }}
.pa-sid  {{ color:var(--text2); font-size:10px }}
.pa-verdict-pills {{ display:flex; gap:5px; margin-left:auto }}
.pa-verdict {{ font-size:10px; font-weight:700; padding:2px 8px; border-radius:4px }}
.pa-yes {{ background:rgba(248,81,73,.15); color:var(--red) }}
.pa-no  {{ background:rgba(139,148,158,.1); color:var(--text2) }}
.pa-rights {{ display:flex; align-items:center; gap:6px; flex-wrap:wrap; font-size:11px }}
.pa-label {{ color:var(--text2); font-size:10px; text-transform:uppercase;
             letter-spacing:.4px; font-weight:700 }}
.perm-pill {{ font-size:10px; padding:2px 7px; border-radius:4px;
              display:inline-block; margin:1px }}
.perm-write  {{ background:rgba(248,81,73,.12); color:var(--red) }}
.perm-enroll {{ background:rgba(88,166,255,.12); color:var(--blue) }}
.perm-read   {{ background:rgba(139,148,158,.1); color:var(--text2) }}
.perm-deny   {{ background:rgba(248,81,73,.08); color:var(--red);
                text-decoration:line-through }}
.pa-missing {{ font-size:11px; color:var(--text2); margin-top:5px;
               background:rgba(63,185,80,.06); border:1px solid rgba(63,185,80,.2);
               border-radius:4px; padding:5px 10px }}
.pa-empty {{ padding:12px; color:var(--text2); font-size:12px; text-align:center }}
/* ── Raw tool output panels ── */
.raw-tools-section {{ margin-top:18px; border:1px solid var(--border);
                      border-radius:8px; overflow:hidden }}
.raw-tools-header {{ background:var(--bg3); padding:10px 16px; display:flex;
                     align-items:center; gap:12px; border-bottom:1px solid var(--border) }}
.raw-tools-title {{ font-size:12px; font-weight:700; color:var(--text) }}
.raw-tools-sub {{ font-size:11px; color:var(--text2) }}
.raw-tools-tabs {{ display:flex; background:var(--bg3);
                   border-bottom:1px solid var(--border) }}
.raw-tab {{ background:transparent; border:none; border-right:1px solid var(--border);
            color:var(--text2); padding:7px 20px; cursor:pointer;
            font-family:var(--font); font-size:11px; font-weight:600;
            transition:all .15s }}
.raw-tab:hover {{ color:var(--text); background:rgba(88,166,255,.06) }}
.raw-tab.active {{ color:var(--cyan); background:rgba(57,211,83,.08);
                   border-bottom:2px solid var(--cyan) }}
.raw-panel {{ padding:0 }}
.active-panel {{ display:block }}
.raw-explain {{ background:rgba(88,166,255,.06); border-bottom:1px solid var(--border);
                padding:10px 16px; font-size:11px; color:var(--text2);
                line-height:1.7 }}
.raw-explain strong {{ color:var(--text) }}
.raw-explain code {{ background:var(--bg3); color:var(--cyan); padding:1px 5px;
                     border-radius:3px; font-size:10px }}
.raw-explain em {{ color:var(--green); font-style:normal; font-weight:600 }}
.raw-output {{ background:var(--bg); color:var(--text); font-family:var(--font);
               font-size:11px; padding:14px 16px; margin:0; overflow-x:auto;
               white-space:pre; line-height:1.6; max-height:400px;
               overflow-y:auto; border:none }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-left">
    <div class="topbar-title">⚡ ADCS Audit Dashboard</div>
    <div class="domain">Domain: {h(domain)} · DC: {h(dc_ip)}</div>
  </div>
  <nav class="topbar-nav">
    <button class="nav-btn active" onclick="showPage('dashboard',this)">📊 Dashboard</button>
    <button class="nav-btn" onclick="showPage('changes',this)">🔄 Changes {diff_badge}</button>
    <button class="nav-btn" onclick="showPage('guide',this)">📖 ESC Guide</button>
  </nav>
  <div class="export-btns">
    <a class="export-btn" href="/export/json" download>⬇ JSON</a>
    <a class="export-btn" href="/export/excel" download>⬇ Excel</a>
    <a class="export-btn export-btn-sec" href="/export/history" download>📁 History</a>
  </div>
</div>

<div id="page-dashboard" class="main">

  <!-- Stats -->
  <div class="stats">
    <div class="stat-card">
      <div class="stat-num num-total">{total}</div>
      <div class="stat-label">Templates</div>
    </div>
    <div class="stat-card">
      <div class="stat-num num-pass">{pub}</div>
      <div class="stat-label">Published</div>
    </div>
    <div class="stat-card">
      <div class="stat-num num-critical">{crits}</div>
      <div class="stat-label">Critical</div>
    </div>
    <div class="stat-card">
      <div class="stat-num num-high">{highs}</div>
      <div class="stat-label">High</div>
    </div>
    <div class="stat-card">
      <div class="stat-num num-medium">{meds}</div>
      <div class="stat-label">Medium</div>
    </div>
    <div class="stat-card">
      <div class="stat-num {'num-critical' if vulns_t else 'num-pass'}">{vulns_t}</div>
      <div class="stat-label">Vulnerable</div>
    </div>
  </div>

  <!-- CAs -->
  <div class="ca-section">
    <h2>Enterprise Certificate Authorities</h2>
    <div class="table-wrap">
      <table>
        <thead><tr><th>CA Name</th><th>Hostname</th><th>ESC Checks</th></tr></thead>
        <tbody>{ca_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- Filter bar -->
  <div class="filter-bar">
    <input type="text" id="search" placeholder="🔍 Search template name…" oninput="applyFilters()">

    <div class="filter-group">
      <span class="filter-group-label">SEVERITY</span>
      <button class="filter-btn sev-btn" data-filter="all" onclick="toggleSev('all',this)">All</button>
      <button class="filter-btn sev-btn" data-filter="critical" onclick="toggleSev('critical',this)">Critical</button>
      <button class="filter-btn sev-btn" data-filter="high" onclick="toggleSev('high',this)">High</button>
      <button class="filter-btn sev-btn" data-filter="medium" onclick="toggleSev('medium',this)">Medium</button>
      <button class="filter-btn sev-btn" data-filter="pass" onclick="toggleSev('pass',this)">Clean</button>
      <button class="filter-btn sev-btn" data-filter="published" onclick="toggleSev('published',this)">Published only</button>
    </div>

    <div class="filter-divider"></div>

    <div class="filter-group">
      <span class="filter-group-label">ESC TYPE</span>
      <button class="filter-btn esc-btn" data-filter="esc1" onclick="toggleEsc('esc1',this)">ESC1</button>
      <button class="filter-btn esc-btn" data-filter="esc1-upn" onclick="toggleEsc('esc1-upn',this)">ESC1&#8209;UPN</button>
      <button class="filter-btn esc-btn" data-filter="esc2" onclick="toggleEsc('esc2',this)">ESC2</button>
      <button class="filter-btn esc-btn" data-filter="esc3" onclick="toggleEsc('esc3',this)">ESC3</button>
      <button class="filter-btn esc-btn" data-filter="esc4" onclick="toggleEsc('esc4',this)">ESC4</button>
      <button class="filter-btn esc-btn" data-filter="esc9" onclick="toggleEsc('esc9',this)">ESC9</button>
    </div>

    <div class="filter-divider"></div>

    <div class="filter-group">
      <span class="filter-group-label">PRINCIPAL TYPE</span>
      <button class="filter-btn tier-btn" data-filter="low" onclick="toggleTier('low',this)">Low Priv Users</button>
      <button class="filter-btn tier-btn" data-filter="unknown" onclick="toggleTier('unknown',this)">Custom Groups</button>
      <button class="filter-btn tier-btn" data-filter="high" onclick="toggleTier('high',this)">Privileged Users</button>
    </div>

    <span class="count-label" id="count-label">{total} templates</span>
  </div>

  <!-- Template cards -->
  <div id="cards-container">
    {cards}
  </div>

  <!-- Legend -->
  <div class="legend">
    <div class="legend-item"><span class="tier-badge tier-low">LOW</span> Everyone, Authenticated Users, Domain Users</div>
    <div class="legend-item"><span class="tier-badge tier-high">HIGH</span> Domain Admins, Enterprise Admins, Administrators</div>
    <div class="legend-item"><span class="tier-badge tier-system">SYSTEM</span> NT AUTHORITY accounts</div>
    <div class="legend-item"><span class="tier-badge tier-unknown">UNKNOWN</span> Custom groups / individual users</div>
    <div class="legend-item" style="color:#8b949e">⚠ Write rights detected on ACE row</div>
  </div>

</div><!-- end page-dashboard -->

<!-- ══════════════════════════════════════════════════════════════════ -->
<!-- ESC GUIDE PAGE                                                     -->
<!-- ══════════════════════════════════════════════════════════════════ -->
<div id="page-guide" class="main" style="display:none">

  <div class="guide-hero">
    <h1 class="guide-hero-title">ADCS ESC Vulnerability Reference</h1>
    <p class="guide-hero-sub">How each attack works, what permissions are needed,
    how to detect false positives, and what makes a finding real.</p>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span class="guide-esc-badge sev-high">ESC4</span>
      <h2>Template Write Access — The Most Common False Positive</h2>
    </div>
    <p class="guide-intro">
      ESC4 triggers when a non-admin principal has <strong>write access</strong> over
      a certificate template object in AD. The key word is <em>write</em> — not every
      permission on an ACE is actually a write to template attributes.
      This is the most common source of false positives in ADCS tooling.
    </p>
    <div class="guide-two-col">
      <div class="guide-col">
        <div class="guide-box box-danger">
          <div class="guide-box-title">⚠ Permissions that ARE exploitable for ESC4</div>
          <table class="guide-table">
            <thead><tr><th>Right</th><th>How</th><th>ESC4?</th></tr></thead>
            <tbody>
              <tr><td><code>WriteDACL</code></td><td>Rewrites template ACL → grants any right</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteOwner</code></td><td>Takes ownership → then grants WriteDACL</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>GenericAll</code></td><td>Includes all rights above</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty</code> (unscoped, no GUID)</td><td>Writes ALL attributes — standard ACE</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty[msPKI-Certificate-Name-Flag]</code></td><td>Directly enables enrollee-supplied SAN → ESC1</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty[msPKI-Enrollment-Flag]</code></td><td>Removes manager approval requirement</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty[pKIExtendedKeyUsage]</code></td><td>Adds Client Auth EKU</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty[msPKI-Certificate-Application-Policy]</code></td><td>Adds Client Auth (takes precedence over EKU)</td><td><span class="yes-badge">YES</span></td></tr>
              <tr><td><code>WriteProperty[msPKI-RA-Signature]</code></td><td>Sets authorized signatures to 0</td><td><span class="yes-badge">YES</span></td></tr>
            </tbody>
          </table>
        </div>
      </div>
      <div class="guide-col">
        <div class="guide-box box-safe">
          <div class="guide-box-title">✓ Permissions that are NOT exploitable for ESC4</div>
          <table class="guide-table">
            <thead><tr><th>Right</th><th>What it means</th><th>ESC4?</th></tr></thead>
            <tbody>
              <tr><td><code>Enroll</code> / Certificate-Enrollment GUID</td><td>Can <em>request</em> a cert — normal for Domain Users</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>AutoEnroll</code> / AutoEnrollment GUID</td><td>Can auto-enroll — normal for workstation templates</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>ReadControl</code></td><td>Read the security descriptor only</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>ReadProperty</code></td><td>Read template attributes only</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>WriteProperty[description]</code></td><td>Description field — irrelevant to cert issuance</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>WriteProperty[userCertificate]</code></td><td>Stored certs — does not affect template logic</td><td><span class="no-badge">NO</span></td></tr>
              <tr><td><code>ListObject</code> / <code>ListContents</code></td><td>Enumerate only — read only</td><td><span class="no-badge">NO</span></td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>
    <div class="guide-box box-info" style="margin-top:16px">
      <div class="guide-box-title">🔑 The Object ACE GUID distinction — why it matters</div>
      <p style="margin-bottom:12px;font-size:13px;color:var(--text2)">
        AD grants <code>WriteProperty</code> in two ways. The ACE type tells you whether
        it covers all attributes or just one specific one:
      </p>
      <div class="ace-compare">
        <div class="ace-box ace-danger">
          <div class="ace-box-title">Standard ACE (type=0x00) — Unscoped ⚠</div>
          <code class="ace-code">(A;;WP;;;S-1-5-21-…-513)</code>
          <p>No GUID. WriteProperty on <strong>ALL attributes</strong>. Always ESC4 if held by non-admin.</p>
          <span class="yes-badge" style="margin-top:8px;display:inline-block">EXPLOITABLE</span>
        </div>
        <div class="ace-arrow">→</div>
        <div class="ace-box ace-safe">
          <div class="ace-box-title">Object ACE (type=0x05) — Scoped ✓</div>
          <code class="ace-code">(OA;;CR;0e10c968-78fb-11d2-90d4-00c04f79dc55;;S-1-5-21-…-513)</code>
          <p>GUID present. WriteProperty/ControlAccess on <strong>one specific attribute</strong>.<br>
          Only ESC4 if the GUID is a security-relevant msPKI attribute.</p>
          <span class="no-badge" style="margin-top:8px;display:inline-block">ENROLL RIGHT ONLY</span>
        </div>
      </div>
    </div>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span class="guide-esc-badge sev-critical">ESC1</span>
      <h2>Enrollee Supplies Subject Alternative Name</h2>
    </div>
    <p class="guide-intro">The most direct path to domain compromise. When a template allows
    the enrollee to specify their own SAN, any permitted user can request a cert for
    <em>administrator@domain</em> and authenticate as Domain Admin.</p>
    <div class="guide-req-grid">
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">CT_FLAG_ENROLLEE_SUPPLIES_SUBJECT</div><div class="req-detail">msPKI-Certificate-Name-Flag bit 0x1</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">No manager approval</div><div class="req-detail">msPKI-Enrollment-Flag bit 0x2 NOT set</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Auth EKU present</div><div class="req-detail">Client Auth / Smartcard / Kerberos / Any Purpose</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Low-priv can enroll</div><div class="req-detail">Domain Users or Authenticated Users has Enroll right</div></div>
    </div>
    <div class="guide-exploit">
      <div class="guide-box-title">Exploit (Certipy v5)</div>
      <code class="exploit-code">certipy req -u USER@DOMAIN -p PASS -dc-ip DC -ca CA_NAME -template TEMPLATE -upn administrator@DOMAIN</code>
    </div>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span class="guide-esc-badge sev-high">ESC2</span>
      <h2>Any Purpose EKU or No EKU Defined</h2>
    </div>
    <p class="guide-intro">A template with no EKUs or the Any Purpose OID issues certs usable
    for anything. Cannot be exploited alone like ESC1 but can be chained as an enrollment
    agent certificate to enroll on behalf of other users.</p>
    <div class="guide-req-grid">
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Any Purpose EKU or no EKU</div><div class="req-detail">pKIExtendedKeyUsage empty or OID 2.5.29.37.0</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">No manager approval</div><div class="req-detail">msPKI-Enrollment-Flag bit 0x2 NOT set</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Low-priv can enroll</div><div class="req-detail">Enroll right for non-admin</div></div>
    </div>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span class="guide-esc-badge sev-high">ESC3</span>
      <h2>Enrollment Agent Certificate</h2>
    </div>
    <p class="guide-intro">A template with the Enrollment Agent EKU allows the holder to request
    certificates on behalf of any other user against a second template. Two-step attack: obtain
    the agent cert, then enroll as Domain Admin on any auth-capable template.</p>
    <div class="guide-req-grid">
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Enrollment Agent EKU</div><div class="req-detail">OID 1.3.6.1.4.1.311.20.2.1</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">No manager approval</div><div class="req-detail">msPKI-Enrollment-Flag bit 0x2 NOT set</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">No RA signature required</div><div class="req-detail">msPKI-RA-Signature = 0</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Low-priv can enroll</div><div class="req-detail">Enroll right on the agent template</div></div>
    </div>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span class="guide-esc-badge sev-high">ESC9</span>
      <h2>No Security Extension (Missing SID Binding)</h2>
    </div>
    <p class="guide-intro">When CT_FLAG_NO_SECURITY_EXTENSION is set, the issued certificate
    omits the szOID_NTDS_CA_SECURITY_EXT extension binding it to the requester's SID.
    Without SID binding, Kerberos strong mapping checks can be bypassed in certain conditions.</p>
    <div class="guide-req-grid">
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">NO_SECURITY_EXTENSION flag</div><div class="req-detail">msPKI-Enrollment-Flag bit 0x00080000 set</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Auth EKU present</div><div class="req-detail">Client Auth or equivalent</div></div>
      <div class="req-card req-must"><div class="req-label">MUST HAVE</div><div class="req-name">Low-priv can enroll</div><div class="req-detail">Enroll right for non-admin</div></div>
      <div class="req-card req-nice"><div class="req-label">WORSENS IMPACT</div><div class="req-name">Enrollee also supplies SAN</div><div class="req-detail">Raises severity from MEDIUM to HIGH</div></div>
    </div>
  </div>

  <div class="guide-section">
    <div class="guide-section-header">
      <span style="font-size:20px">🎯</span>
      <h2>How Severity is Assigned</h2>
    </div>
    <div class="guide-sev-grid">
      <div class="sev-explain sev-explain-critical">
        <div class="sev-explain-title"><span class="sev-dot-lg" style="background:var(--red)"></span>CRITICAL</div>
        <p>Direct domain takeover. Enrollee controls the SAN and can impersonate Domain Admins. No additional steps required.</p>
        <p><strong>ESC1, ESC1-UPN</strong></p>
      </div>
      <div class="sev-explain sev-explain-high">
        <div class="sev-explain-title"><span class="sev-dot-lg sev-high"></span>HIGH</div>
        <p>Exploitation requires 1–2 extra steps (modify template, get agent cert, chain attacks) but leads to full compromise.</p>
        <p><strong>ESC2, ESC3, ESC4, ESC9 (with SAN)</strong></p>
      </div>
      <div class="sev-explain sev-explain-medium">
        <div class="sev-explain-title"><span class="sev-dot-lg sev-medium"></span>MEDIUM</div>
        <p>Requires additional conditions outside the template (shadow credentials, NTLM relay) to exploit.</p>
        <p><strong>ESC9 (without SAN)</strong></p>
      </div>
      <div class="sev-explain sev-explain-pass">
        <div class="sev-explain-title"><span class="sev-dot-lg sev-pass"></span>PASS / CLEAN</div>
        <p>No vulnerability confirmed. Does NOT mean safe — check individual condition counts for residual risk.</p>
        <p><strong>All ESC checks fail (at least one condition missing)</strong></p>
      </div>
    </div>
  </div>

</div><!-- end page-guide -->

<!-- ── CHANGES PAGE ── -->
<div id="page-changes" class="main" style="display:none">
  <div class="guide-hero">
    <h1 class="guide-hero-title">🔄 Changes vs Previous Scan</h1>
    <p class="guide-hero-sub">{changes_summary_html}</p>
  </div>
  {changes_content_html}
</div><!-- end page-changes -->

<script>
function switchRawTab(btn, panelId) {{
  var section = btn.closest('.raw-tools-section');
  section.querySelectorAll('.raw-tab').forEach(b => b.classList.remove('active'));
  section.querySelectorAll('.raw-panel').forEach(p => p.style.display='none');
  btn.classList.add('active');
  var panel = document.getElementById(panelId);
  if (panel) panel.style.display='block';
}}

function showPage(name, btn) {{
  document.querySelectorAll('[id^="page-"]').forEach(p => p.style.display='none');
  document.getElementById('page-'+name).style.display='block';
  document.querySelectorAll('.nav-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
}}

// ── Multi-select filter state ─────────────────────────────────────
var activeSev   = new Set();   // empty = show all
var activeEsc   = new Set();   // empty = show all
var activeTier  = new Set();   // empty = show all

function toggleCard(hdr) {{
  var body = hdr.nextElementSibling;
  var chev = hdr.querySelector('.chevron');
  var open = body.style.display !== 'none';
  body.style.display = open ? 'none' : 'block';
  chev.classList.toggle('open', !open);
}}

function toggleSev(f, btn) {{
  if (f === 'all') {{
    activeSev.clear();
    document.querySelectorAll('.sev-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
  }} else {{
    // deselect "All" when picking a specific severity
    document.querySelector('.sev-btn[data-filter="all"]').classList.remove('active');
    if (activeSev.has(f)) {{
      activeSev.delete(f); btn.classList.remove('active');
      // if nothing selected, re-activate All
      if (activeSev.size === 0)
        document.querySelector('.sev-btn[data-filter="all"]').classList.add('active');
    }} else {{
      activeSev.add(f); btn.classList.add('active');
    }}
  }}
  applyFilters();
}}

function toggleEsc(f, btn) {{
  if (activeEsc.has(f)) {{
    activeEsc.delete(f); btn.classList.remove('active');
  }} else {{
    activeEsc.add(f); btn.classList.add('active');
  }}
  applyFilters();
}}

function toggleTier(f, btn) {{
  if (activeTier.has(f)) {{
    activeTier.delete(f); btn.classList.remove('active');
  }} else {{
    activeTier.add(f); btn.classList.add('active');
  }}
  applyFilters();
}}

function applyFilters() {{
  var q     = document.getElementById('search').value.toLowerCase();
  var cards = document.querySelectorAll('.template-card');
  var shown = 0;

  cards.forEach(function(c) {{
    var sev   = c.dataset.sev;
    var pub   = c.dataset.published === 'true';
    var name  = c.dataset.name;
    var escs  = (c.dataset.escs  || '').split(',').filter(Boolean);
    var tiers = (c.dataset.tiers || '').split(',').filter(Boolean);

    // ── Severity group (OR within group)
    var sevOk = activeSev.size === 0 ||
                activeSev.has(sev) ||
                (activeSev.has('published') && pub);

    // ── ESC group (OR within group)
    var escOk = activeEsc.size === 0 ||
                escs.some(e => activeEsc.has(e));

    // ── Tier/principal group (OR within group)
    var tierOk = activeTier.size === 0 ||
                 tiers.some(t => activeTier.has(t));

    // ── Search
    var srchOk = !q || name.includes(q);

    // ── All groups AND together
    var vis = sevOk && escOk && tierOk && srchOk;
    c.style.display = vis ? '' : 'none';
    if (vis) shown++;
  }});

  document.getElementById('count-label').textContent =
    shown + ' template' + (shown !== 1 ? 's' : '') +
    (activeEsc.size || activeTier.size || activeSev.size ?
      ' <span style="color:var(--blue)">(filtered)</span>' : '');
}}

// Auto-expand vulnerable cards on load
document.addEventListener('DOMContentLoaded', function() {{
  document.querySelectorAll('.template-card').forEach(function(c) {{
    if (c.dataset.sev !== 'pass') {{
      var body = c.querySelector('.card-body');
      var chev = c.querySelector('.chevron');
      body.style.display = 'block';
      chev.classList.add('open');
    }}
  }});
  // start with "All" active
  document.querySelector('.sev-btn[data-filter="all"]').classList.add('active');
}});
</script>
</body>
</html>"""

# ══════════════════════════════════════════════════════════
# HTTP SERVER
# ══════════════════════════════════════════════════════════

_HTML_CONTENT  = ""
_SCAN_RESULTS  = {}   # {"ca_info": [...], "results": [...], "domain": "", "dc_ip": "", "ts": ""}
_HISTORY_FILE  = "adcs_history.json"
_HISTORY       = {}   # {scan_hash: scan_snapshot}

# ══════════════════════════════════════════════════════════
# HISTORY / DIFF ENGINE
# ══════════════════════════════════════════════════════════

def results_to_snapshot(ca_info, results, domain, dc_ip):
    """Serialise results into a JSON-safe dict for history storage."""
    ts = datetime.datetime.now().isoformat(timespec="seconds")
    snap = {
        "ts": ts,
        "domain": domain,
        "dc_ip": dc_ip,
        "cas": [{"cn": c["cn"], "dns": c["dns"]} for c in ca_info],
        "templates": []
    }
    for r in results:
        snap["templates"].append({
            "name":        r["name"],
            "published":   r["published"],
            "cas":         r["cas"],
            "max_sev":     r["max_sev"],
            "san_flag":    r["san_flag"],
            "upn_flag":    r["upn_flag"],
            "approval":    r["approval"],
            "no_sec":      r["no_sec"],
            "schema":      r["schema"],
            "min_key":     r["min_key"],
            "ra":          r["ra"],
            "eku":         r["eku"],
            "acl": [{"name": e["name"], "tier": e["tier"],
                     "rights": e["rights"], "deny": e["deny"],
                     "dangerous_write": e.get("dangerous_write", False),
                     "write_reason": e.get("write_reason","")} for e in r["acl"]],
            "vulns": [{"esc": v["esc"], "sev": v["sev"]} for v in r["vulns"]],
        })
    snap["hash"] = hashlib.md5(
        json.dumps(snap["templates"], sort_keys=True).encode()
    ).hexdigest()[:12]
    return snap

def load_history():
    global _HISTORY
    if os.path.exists(_HISTORY_FILE):
        try:
            with open(_HISTORY_FILE, "r", encoding="utf-8") as f:
                _HISTORY = json.load(f)
            print(f"[*] Loaded {len(_HISTORY)} historical scan(s) from {_HISTORY_FILE}")
        except Exception as e:
            print(f"[!] Could not load history: {e}")
            _HISTORY = {}

def save_history(snap):
    global _HISTORY
    _HISTORY[snap["hash"]] = snap
    # keep only last 20 scans
    if len(_HISTORY) > 20:
        oldest = sorted(_HISTORY.values(), key=lambda x: x["ts"])[0]["hash"]
        del _HISTORY[oldest]
    try:
        with open(_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(_HISTORY, f, indent=2)
        print(f"[+] Scan snapshot saved to {_HISTORY_FILE} (hash={snap['hash']})")
    except Exception as e:
        print(f"[!] Could not save history: {e}")

def diff_scans(prev_snap, curr_snap):
    """Compare two snapshots and return list of change dicts."""
    prev = {t["name"]: t for t in prev_snap["templates"]}
    curr = {t["name"]: t for t in curr_snap["templates"]}
    changes = []

    # New templates
    for name in curr:
        if name not in prev:
            changes.append({"type":"new_template","template":name,
                            "sev":curr[name]["max_sev"],
                            "detail":"Template did not exist in previous scan"})

    # Removed templates
    for name in prev:
        if name not in curr:
            changes.append({"type":"removed_template","template":name,
                            "detail":"Template no longer present"})

    # Changed templates
    for name in curr:
        if name not in prev:
            continue
        c = curr[name]; p = prev[name]

        # Severity change
        sev_order = {"critical":0,"high":1,"medium":2,"pass":3}
        if c["max_sev"] != p["max_sev"]:
            direction = "WORSENED" if sev_order[c["max_sev"]] < sev_order[p["max_sev"]] else "IMPROVED"
            changes.append({"type":"severity_change","template":name,
                            "direction":direction,
                            "old_sev":p["max_sev"],"new_sev":c["max_sev"],
                            "detail":f"Severity {direction}: {p['max_sev'].upper()} → {c['max_sev'].upper()}"})

        # New vulnerabilities
        prev_escs = {v["esc"] for v in p["vulns"]}
        curr_escs = {v["esc"] for v in c["vulns"]}
        for esc in curr_escs - prev_escs:
            changes.append({"type":"new_vuln","template":name,"esc":esc,
                            "detail":f"New {esc} vulnerability appeared"})
        for esc in prev_escs - curr_escs:
            changes.append({"type":"fixed_vuln","template":name,"esc":esc,
                            "detail":f"{esc} vulnerability was remediated"})

        # ACL changes
        prev_acl = {(e["name"],e["deny"]): set(e["rights"]) for e in p["acl"]}
        curr_acl = {(e["name"],e["deny"]): set(e["rights"]) for e in c["acl"]}
        for key in curr_acl:
            if key not in prev_acl:
                pname, deny = key
                changes.append({"type":"acl_added","template":name,
                                "principal":pname,"deny":deny,
                                "rights":list(curr_acl[key]),
                                "detail":f"New ACE added for {pname}"})
        for key in prev_acl:
            if key not in curr_acl:
                pname, deny = key
                changes.append({"type":"acl_removed","template":name,
                                "principal":pname,
                                "detail":f"ACE removed for {pname}"})
            else:
                gained = curr_acl[key] - prev_acl[key]
                lost   = prev_acl[key] - curr_acl[key]
                if gained:
                    pname, deny = key
                    changes.append({"type":"rights_gained","template":name,
                                    "principal":pname,"rights":list(gained),
                                    "detail":f"{pname} gained rights: {', '.join(gained)}"})
                if lost:
                    pname, deny = key
                    changes.append({"type":"rights_lost","template":name,
                                    "principal":pname,"rights":list(lost),
                                    "detail":f"{pname} lost rights: {', '.join(lost)}"})

        # Published status
        if c["published"] != p["published"]:
            changes.append({"type":"published_change","template":name,
                            "detail":f"Published status: {p['published']} → {c['published']}"})

    return changes

# ══════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════

def build_excel(ca_info, results, domain, diff_changes=None):
    """Build an xlsx file in memory, return bytes."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # ── Color palette ───────────────────────────────────────
    RED    = PatternFill("solid", fgColor="F85149")
    ORANGE = PatternFill("solid", fgColor="E3B341")
    YELLOW = PatternFill("solid", fgColor="D29922")
    GREEN  = PatternFill("solid", fgColor="3FB950")
    GREY   = PatternFill("solid", fgColor="30363D")
    BLUE   = PatternFill("solid", fgColor="1F6FEB")
    DARK   = PatternFill("solid", fgColor="161B22")
    HDR_BG = PatternFill("solid", fgColor="21262D")
    WHITE  = Font(color="E6EDF3", bold=True)
    NORMAL = Font(color="E6EDF3")
    SEV_FILL = {"critical":RED,"high":ORANGE,"medium":YELLOW,"pass":GREEN}
    TIER_FILL = {"high":ORANGE,"low":RED,"unknown":GREY,"system":BLUE}

    def hdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = HDR_BG; cell.font = WHITE
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def thin_border():
        s = Side(style="thin", color="30363D")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Summary ─────────────────────────────────────
    ws1 = wb.active; ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "F85149"
    hdr(ws1, 1, ["Template","Published","CAs","Severity",
                  "ESC Findings","SAN Flag","Approval","Schema","Min Key","EKUs"])
    set_col_widths(ws1, [30, 10, 30, 12, 20, 10, 10, 8, 8, 50])
    for r in results:
        row = [
            r["name"],
            "YES" if r["published"] else "NO",
            ", ".join(r["cas"]) or "—",
            r["max_sev"].upper(),
            ", ".join(v["esc"] for v in r["vulns"]) or "—",
            "YES" if r["san_flag"] else "NO",
            "REQUIRED" if r["approval"] else "NO",
            r["schema"],
            r["min_key"],
            ", ".join(r["eku"]) or "—",
        ]
        ws1.append(row)
        sev_cell = ws1.cell(ws1.max_row, 4)
        sev_cell.fill = SEV_FILL.get(r["max_sev"], GREY)
        sev_cell.font = Font(color="0D1117", bold=True)
        for c in range(1, 11):
            ws1.cell(ws1.max_row, c).border = thin_border()

    # ── Sheet 2: ACL Details ─────────────────────────────────
    ws2 = wb.create_sheet("ACL Details")
    ws2.sheet_view.showGridLines = False
    hdr(ws2, 1, ["Template","Principal","SID","Tier","Allow/Deny",
                  "Rights","Dangerous Write?","Write Reason","ESC4?"])
    set_col_widths(ws2, [28,28,35,10,10,40,14,40,8])
    for r in results:
        for e in r["acl"]:
            row = [
                r["name"],
                e["name"],
                e.get("sid",""),
                e["tier"].upper(),
                "DENY" if e["deny"] else "ALLOW",
                ", ".join(e["rights"]),
                "YES" if e.get("dangerous_write") else "NO",
                e.get("write_reason",""),
                "YES" if (e.get("dangerous_write") and e["tier"] not in ("high","system")) else "NO",
            ]
            ws2.append(row)
            row_n = ws2.max_row
            ws2.cell(row_n,4).fill = TIER_FILL.get(e["tier"],GREY)
            ws2.cell(row_n,4).font = Font(color="0D1117",bold=True)
            if e.get("dangerous_write") and e["tier"] not in ("high","system"):
                ws2.cell(row_n,9).fill = RED
                ws2.cell(row_n,9).font = Font(color="0D1117",bold=True)
            for c in range(1,10):
                ws2.cell(row_n,c).border = thin_border()

    # ── Sheet 3: ESC Findings ────────────────────────────────
    ws3 = wb.create_sheet("ESC Findings")
    ws3.sheet_view.showGridLines = False
    hdr(ws3, 1, ["Template","Published On","ESC","Severity",
                  "All Conditions Met","Triggering Principals"])
    set_col_widths(ws3, [28,35,14,12,18,50])
    for r in results:
        for chk in r.get("checks", []):
            if not chk["vuln"]: continue
            princs = "; ".join(p[0] for p in chk.get("principals",[])[:5])
            row = [
                r["name"],
                ", ".join(r["cas"]) or "not published",
                chk["esc"],
                chk["sev"].upper(),
                "YES",
                princs or "—",
            ]
            ws3.append(row)
            row_n = ws3.max_row
            ws3.cell(row_n,4).fill = SEV_FILL.get(chk["sev"],GREY)
            ws3.cell(row_n,4).font = Font(color="0D1117",bold=True)
            for c in range(1,7):
                ws3.cell(row_n,c).border = thin_border()

    # ── Sheet 4: Diff / Changes ──────────────────────────────
    if diff_changes:
        ws4 = wb.create_sheet("Changes vs Previous")
        ws4.sheet_view.showGridLines = False
        ws4.sheet_properties.tabColor = "E3B341"
        hdr(ws4, 1, ["Change Type","Template","Detail","Principal","Rights","Direction"])
        set_col_widths(ws4, [20,28,50,25,30,12])
        for ch in diff_changes:
            row = [
                ch["type"].replace("_"," ").upper(),
                ch.get("template",""),
                ch.get("detail",""),
                ch.get("principal",""),
                ", ".join(ch.get("rights",[])),
                ch.get("direction",""),
            ]
            ws4.append(row)
            row_n = ws4.max_row
            t = ch["type"]
            fill = RED if "worsened" in ch.get("direction","").lower() or "new_vuln" in t or "gained" in t                    else GREEN if "improved" in ch.get("direction","").lower() or "fixed" in t or "lost" in t                    else ORANGE
            ws4.cell(row_n,1).fill = fill
            ws4.cell(row_n,1).font = Font(color="0D1117",bold=True)
            for c in range(1,7):
                ws4.cell(row_n,c).border = thin_border()

    # ── Sheet 5: CAs ─────────────────────────────────────────
    ws5 = wb.create_sheet("Certificate Authorities")
    ws5.sheet_view.showGridLines = False
    hdr(ws5, 1, ["CA Name","DNS Hostname","ESC8 Check URL"])
    set_col_widths(ws5, [40,40,50])
    for ca in ca_info:
        ws5.append([ca["cn"], ca["dns"], f"http://{ca['dns']}/certsrv/certfnsh.asp"])
        for c in range(1,4):
            ws5.cell(ws5.max_row,c).border = thin_border()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_template_attrs_output(tname, dn, t_raw, nf, ef, ra, eku_list):
    """
    Tab: Template Attributes
    Relevant for: ESC1, ESC1-UPN, ESC2, ESC3, ESC9
    Shows ldapsearch for msPKI-* attributes with decoded flag analysis.
    """
    lines = []
    lines.append("# ── RELEVANT FOR: ESC1, ESC1-UPN, ESC2, ESC3, ESC9 ──")
    lines.append("")
    lines.append("# Command (fetches template configuration attributes):")
    lines.append("ldapsearch -x -H ldap://DC_IP -D 'USER@DOMAIN' -w 'PASS'")
    lines.append(f"  -b '{dn}'")
    lines.append("  '(objectClass=*)' cn msPKI-Certificate-Name-Flag msPKI-Enrollment-Flag")
    lines.append("  msPKI-RA-Signature pKIExtendedKeyUsage msPKI-Certificate-Application-Policy")
    lines.append("  msPKI-Template-Schema-Version msPKI-Minimal-Key-Size")
    lines.append("")
    lines.append(f"# Result for template: {tname}")
    lines.append(f"dn: {dn}")
    lines.append(f"cn: {tname}")
    lines.append(f"msPKI-Certificate-Name-Flag: {nf} (0x{nf:08x})")
    lines.append(f"msPKI-Enrollment-Flag:        {ef} (0x{ef:08x})")
    lines.append(f"msPKI-RA-Signature:           {ra}")
    for e in (eku_list or []):
        lines.append(f"pKIExtendedKeyUsage: {e}")
    lines.append("")
    lines.append("# ── DECODED FLAG ANALYSIS ──")
    lines.append("")

    # msPKI-Certificate-Name-Flag
    lines.append(f"msPKI-Certificate-Name-Flag = 0x{nf:08x}")
    CT_FLAGS_NAME = {
        0x00000001: ("CT_FLAG_ENROLLEE_SUPPLIES_SUBJECT",    "ESC1 ← DANGEROUS if auth EKU present"),
        0x00000002: ("CT_FLAG_ENROLLEE_SUPPLIES_SUBJECT_ALT_NAME","ESC1-UPN ← DANGEROUS"),
        0x00000008: ("CT_FLAG_OLD_CERT_SUPPLIES_SUBJECT",    ""),
        0x00000040: ("CT_FLAG_SUBJECT_REQUIRE_DN",           ""),
        0x00400000: ("CT_FLAG_SUBJECT_ALT_REQUIRE_DOMAIN_DNS",""),
        0x02000000: ("CT_FLAG_SUBJECT_ALT_REQUIRE_UPN",      "normal for user templates"),
    }
    for bit, (fname, note) in CT_FLAGS_NAME.items():
        val = "SET" if nf & bit else "not set"
        marker = " *** " if ("ESC" in note and nf & bit) else "     "
        lines.append(f"  {marker}bit 0x{bit:08x} [{fname}]: {val}  {note}")

    lines.append("")
    lines.append(f"msPKI-Enrollment-Flag = 0x{ef:08x}")
    EF_FLAGS = {
        0x00000002: ("CT_FLAG_PEND_ALL_REQUESTS",     "Manager approval REQUIRED — mitigates ESC1/2/3"),
        0x00000001: ("CT_FLAG_INCLUDE_SYMMETRIC_ALGORITHMS",""),
        0x00000020: ("CT_FLAG_AUTO_ENROLLMENT",        ""),
        0x00080000: ("CT_FLAG_NO_SECURITY_EXTENSION",  "ESC9 ← missing SID binding in cert"),
    }
    for bit, (fname, note) in EF_FLAGS.items():
        val = "SET" if ef & bit else "not set"
        marker = " *** " if ("ESC" in note and ef & bit) else "     "
        lines.append(f"  {marker}bit 0x{bit:08x} [{fname}]: {val}  {note}")

    lines.append("")
    lines.append(f"msPKI-RA-Signature = {ra}")
    lines.append(f"  {'SAFE — RA signature required before enrollment' if ra > 0 else '  0 = no authorized signature required (ESC3 condition met)'}")

    lines.append("")
    lines.append("# ── EKU ANALYSIS ──")
    AUTH_EKUS_MAP = {
        "1.3.6.1.5.5.7.3.2":       "Client Authentication ← AUTH EKU (ESC1/2/3 condition met)",
        "1.3.6.1.4.1.311.20.2.2":  "Smartcard Logon       ← AUTH EKU",
        "1.3.6.1.5.2.3.5":         "Kerberos Auth         ← AUTH EKU",
        "1.3.6.1.5.2.3.4":         "PKINIT Client         ← AUTH EKU",
        "2.5.29.37.0":              "Any Purpose *** ESC2 condition met",
        "1.3.6.1.4.1.311.20.2.1":  "Enrollment Agent *** ESC3 condition met",
        "1.3.6.1.5.5.7.3.1":       "Server Authentication (not an auth EKU for PKINIT)",
        "1.3.6.1.5.5.7.3.4":       "Email Protection",
        "1.3.6.1.5.5.7.3.9":       "OCSP Signing",
    }
    if not eku_list:
        lines.append("  (no EKUs defined) *** ESC2 condition met — Any Purpose implied")
    else:
        for e in eku_list:
            desc = AUTH_EKUS_MAP.get(e, e)
            marker = " *** " if "ESC" in desc or "AUTH" in desc else "     "
            lines.append(f"  {marker}{e}  =  {desc}")

    lines.append("")
    lines.append("# ── ESC1/2/3/9 VERDICT ──")
    c_san = bool(nf & 0x1); c_upn = bool(nf & 0x2)
    c_approval = bool(ef & 0x2); c_nosec = bool(ef & 0x00080000)
    auth_eku = not eku_list or any(e in AUTH_EKUS for e in eku_list)
    any_eku  = not eku_list or "2.5.29.37.0" in eku_list
    ea_eku   = "1.3.6.1.4.1.311.20.2.1" in (eku_list or [])
    lines.append(f"  ESC1  : {'CONDITION MET' if c_san else 'SAFE'} (Enrollee supplies SAN = {c_san})")
    lines.append(f"  ESC1u : {'CONDITION MET' if c_upn else 'SAFE'} (Enrollee supplies UPN = {c_upn})")
    lines.append(f"  ESC2  : {'CONDITION MET' if any_eku else 'SAFE'} (Any Purpose/no EKU = {any_eku})")
    lines.append(f"  ESC3  : {'CONDITION MET' if ea_eku else 'SAFE'} (Enrollment Agent EKU = {ea_eku})")
    lines.append(f"  ESC9  : {'CONDITION MET' if c_nosec else 'SAFE'} (NO_SECURITY_EXTENSION = {c_nosec})")
    lines.append(f"  Approval required: {c_approval}  (mitigates ESC1/2/3 if True)")
    return "\n".join(lines)


def build_ca_checks_output(tname, ca_info, dc_ip, domain):
    """
    Tab: CA-Level Checks
    Relevant for: ESC5, ESC6, ESC7, ESC8
    Shows commands and status for CA-level vulnerabilities.
    """
    lines = []
    lines.append("# ── RELEVANT FOR: ESC5, ESC6, ESC7, ESC8 ──")
    lines.append("")
    lines.append("# These checks are at the CA level, not per-template.")
    lines.append("# Run these commands to verify each ESC independently.")
    lines.append("")

    # ESC5
    lines.append("# ════ ESC5: PKI Container Object ACLs ════")
    lines.append("# Non-admin write on CN=Public Key Services or child containers")
    lines.append("# allows adding/modifying templates, CAs, or trust anchors.")
    lines.append("")
    containers = [
        "CN=Public Key Services,CN=Services,CN=Configuration,DC=DOMAIN,DC=NET",
        "CN=Certificate Templates,CN=Public Key Services,CN=Services,CN=Configuration,DC=DOMAIN,DC=NET",
        "CN=NTAuthCertificates,CN=Public Key Services,CN=Services,CN=Configuration,DC=DOMAIN,DC=NET",
    ]
    for c in containers:
        lines.append(f"ldapsearch -x -H ldap://DC_IP -D 'USER@DOMAIN' -w 'PASS'")
        lines.append(f"  -b '{c}' '(objectClass=*)' nTSecurityDescriptor")
        lines.append(f"  -E '!1.2.840.113556.1.4.801=::MAMCAQc='")
        lines.append("")

    # ESC6
    lines.append("# ════ ESC6: EDITF_ATTRIBUTESUBJECTALTNAME2 on CA ════")
    lines.append("# If set, the CA accepts SAN from the CSR on ANY template.")
    lines.append("# Cannot be read via LDAP — requires CA server access.")
    lines.append("")
    for ca in (ca_info or []):
        cn  = ca.get("cn","CA")
        dns = ca.get("dns","CA_HOST")
        lines.append(f"# CA: {cn}")
        lines.append("certutil -config '" + dns + "\\" + cn + "' -getreg policy\\EditFlags")
        lines.append(f"  Look for: EDITF_ATTRIBUTESUBJECTALTNAME2 (bit 0x00040000)")
        lines.append(f"  If set: ANY enrolled cert can have arbitrary SAN = all templates vulnerable to ESC1")
        lines.append("")

    # ESC7
    lines.append("# ════ ESC7: ManageCA / ManageCertificates on CA ════")
    lines.append("# Low-priv principal with these rights can approve pending requests")
    lines.append("# or enable EDITF_ATTRIBUTESUBJECTALTNAME2.")
    lines.append("")
    for ca in (ca_info or []):
        cn  = ca.get("cn","CA")
        dns = ca.get("dns","CA_HOST")
        base_domain = ".".join(dns.split(".")[-2:]) if dns != "CA_HOST" else domain
        lines.append(f"# CA: {cn}")
        esc7_vuln = ca.get("esc7_vuln", False)
        lines.append(f"# ESC7 status: {'*** VULNERABLE — non-admin has CA management rights ***' if esc7_vuln else 'SAFE — no low-priv CA management rights detected'}")
        lines.append(f"impacket-dacledit -action read")
        lines.append(f"  -target-dn 'CN={cn},CN=Enrollment Services,CN=Public Key Services,CN=Services,CN=Configuration,DC=...'")
        lines.append(f"  'DOMAIN/USER:PASS' -dc-ip {dc_ip}")
        lines.append("")
        lines.append(f"bloodyAD -u USER -p PASS -d {domain} --host {dc_ip}")
        lines.append(f"  get object 'CN={cn},CN=Enrollment Services,...' --attr nTSecurityDescriptor")
        lines.append("")

    # ESC8
    lines.append("# ════ ESC8: Web Enrollment NTLM Relay ════")
    lines.append("# If web enrollment is accessible over HTTP (not HTTPS), an attacker")
    lines.append("# can relay NTLM auth to the CA and obtain a cert as any machine.")
    lines.append("")
    for ca in (ca_info or []):
        dns = ca.get("dns","CA_HOST")
        esc8 = ca.get("esc8_accessible", False)
        lines.append(f"# CA: {ca.get('cn','?')} ({dns})")
        lines.append(f"# ESC8 status: {'*** ACCESSIBLE — HTTP endpoint reachable ***' if esc8 else 'Not verified (manual check required)'}")
        lines.append(f"curl -v -s -o /dev/null -w '%{{http_code}}' --max-time 5 http://{dns}/certsrv/certfnsh.asp")
        lines.append(f"  401 + WWW-Authenticate: NTLM = VULNERABLE to ESC8")
        lines.append(f"  301/302 to HTTPS        = Protected by channel binding")
        lines.append(f"  Connection refused      = Web enrollment not installed")
        lines.append("")
        lines.append(f"# Full NTLM relay check:")
        lines.append(f"curl -v --ntlm -u 'USER:PASS' http://{dns}/certsrv/certfnsh.asp")
        lines.append("")

    return "\n".join(lines)


def build_json_export(ca_info, results, domain, dc_ip, diff_changes=None):
    """Build a clean JSON export."""
    export = {
        "meta": {
            "domain": domain,
            "dc_ip": dc_ip,
            "timestamp": datetime.datetime.now().isoformat(),
            "tool": "ADCS Web Audit v1.0"
        },
        "certificate_authorities": ca_info,
        "templates": [],
        "findings_summary": {
            "total": len(results),
            "critical": sum(1 for r in results if r["max_sev"]=="critical"),
            "high":     sum(1 for r in results if r["max_sev"]=="high"),
            "medium":   sum(1 for r in results if r["max_sev"]=="medium"),
            "pass":     sum(1 for r in results if r["max_sev"]=="pass"),
        },
        "changes_vs_previous": diff_changes or []
    }
    for r in results:
        export["templates"].append({
            "name": r["name"],
            "display_name": r["display"],
            "published": r["published"],
            "published_on": r["cas"],
            "severity": r["max_sev"],
            "schema_version": r["schema"],
            "min_key_size": r["min_key"],
            "flags": {
                "enrollee_supplies_san": r["san_flag"],
                "enrollee_supplies_upn": r["upn_flag"],
                "manager_approval_required": r["approval"],
                "no_security_extension": r["no_sec"],
                "ra_signatures_required": r["ra"],
            },
            "eku": r["eku"],
            "acl": [{"principal": e["name"], "sid": e.get("sid",""),
                     "tier": e["tier"], "action": "deny" if e["deny"] else "allow",
                     "rights": e["rights"],
                     "dangerous_write": e.get("dangerous_write",False),
                     "write_reason": e.get("write_reason","")} for e in r["acl"]],
            "esc_checks": [{"esc": c["esc"], "severity": c["sev"],
                            "vulnerable": c["vuln"],
                            "conditions": {k: v for k,v in c.get("conds",[])},
                            "triggering_principals": [{"name":p[0],"rights":p[1],"tier":p[2]}
                                                      for p in c.get("principals",[])]}
                           for c in r.get("checks",[])],
        })
    return json.dumps(export, indent=2)

# ══════════════════════════════════════════════════════════
# HTTP SERVER
# ══════════════════════════════════════════════════════════

_HTML_CONTENT = ""

class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?")[0]

        if path == "/export/json":
            snap   = _SCAN_RESULTS
            diff   = _SCAN_RESULTS.get("diff_changes", [])
            data   = build_json_export(snap["ca_info"], snap["results"],
                                       snap["domain"], snap["dc_ip"], diff)
            fname  = f"adcs_audit_{snap['domain']}_{snap['ts'].replace(':','-')[:19]}.json"
            self._send(data.encode(), "application/json", fname)

        elif path == "/export/excel":
            snap   = _SCAN_RESULTS
            diff   = _SCAN_RESULTS.get("diff_changes", [])
            data   = build_excel(snap["ca_info"], snap["results"],
                                 snap["domain"], diff)
            if data is None:
                self._send(b"openpyxl not installed. Run: pip install openpyxl",
                           "text/plain")
            else:
                fname = f"adcs_audit_{snap['domain']}_{snap['ts'].replace(':','-')[:19]}.xlsx"
                self._send(data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname)

        elif path == "/export/history":
            data = json.dumps(list(_HISTORY.values()), indent=2).encode()
            self._send(data, "application/json", "adcs_history.json")

        elif path == "/api/history":
            # Return history list as JSON for the dashboard
            hist_list = sorted(_HISTORY.values(), key=lambda x: x["ts"], reverse=True)
            payload = [{"hash":s["hash"],"ts":s["ts"],"domain":s["domain"],
                        "templates":len(s["templates"])} for s in hist_list]
            self._send(json.dumps(payload).encode(), "application/json")

        else:
            self._send(_HTML_CONTENT.encode(), "text/html; charset=utf-8")

    def _send(self, data, content_type, filename=None):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        print(f"  [{self.address_string()}] {fmt % args}")

# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

def main():
    global _HTML_CONTENT

    ap = argparse.ArgumentParser(description="ADCS Web Audit — serves on http://0.0.0.0:4000")
    ap.add_argument("-u","--user",     required=True)
    ap.add_argument("-p","--password", required=True)
    ap.add_argument("-d","--dc-ip",    required=True)
    ap.add_argument("--domain",        help="FQDN — auto-detected from -u if omitted")
    ap.add_argument("--no-ntlm",       action="store_true")
    ap.add_argument("--port",          type=int, default=4000)
    ap.add_argument("--output",        metavar="FILE",
                    help="Also save HTML to file (e.g. report.html)")
    args = ap.parse_args()

    domain = args.domain
    if not domain:
        u = args.user
        if "@"  in u: domain = u.split("@")[1]
        elif "\\" in u: domain = u.split("\\")[0]
        else: print("[!] Pass --domain"); sys.exit(1)

    base = ",".join(f"DC={p}" for p in domain.split("."))

    print(f"\n[*] Domain : {domain}")
    print(f"[*] Base DN: {base}")
    print(f"[*] Connecting to ldap://{args.dc_ip}:389 ...")

    try:
        srv  = Server(args.dc_ip, port=389, use_ssl=False,
                      get_info=ALL, connect_timeout=30)
        if not args.no_ntlm:
            u    = f"{domain}\\{args.user.split('@')[0]}"
            conn = Connection(srv, user=u, password=args.password,
                              authentication=NTLM, auto_bind=True, receive_timeout=60)
        else:
            conn = Connection(srv, user=args.user, password=args.password,
                              authentication=SIMPLE, auto_bind=True, receive_timeout=60)
        print("[+] Authenticated")
    except LDAPBindError as e:
        print(f"[!] Auth failed: {e}"); sys.exit(1)
    except Exception as e:
        print(f"[!] Connection error: {e}"); sys.exit(1)

    load_history()

    print("[*] Collecting data (this may take 30-60s for large domains) ...")
    ca_info, results = collect(conn, base)
    ts = datetime.datetime.now().isoformat(timespec="seconds")
    print(f"[+] {len(ca_info)} CAs, {len(results)} templates, {len(_SC)} SIDs resolved")

    # Build snapshot and diff
    curr_snap = results_to_snapshot(ca_info, results, domain, args.dc_ip)
    diff_changes = []
    prev_snap = None
    if _HISTORY:
        prev_snap = sorted(_HISTORY.values(), key=lambda x: x["ts"])[-1]
        if prev_snap["hash"] != curr_snap["hash"]:
            diff_changes = diff_scans(prev_snap, curr_snap)
            print(f"[+] {len(diff_changes)} change(s) detected vs previous scan ({prev_snap['ts']})")
        else:
            print(f"[*] No changes detected vs previous scan ({prev_snap['ts']})")
    else:
        print("[*] No previous scan history — this is the baseline")

    save_history(curr_snap)

    # Store globally for export endpoints
    _SCAN_RESULTS.update({
        "ca_info": ca_info, "results": results,
        "domain": domain, "dc_ip": args.dc_ip,
        "ts": ts, "diff_changes": diff_changes,
        "prev_snap": prev_snap, "curr_snap": curr_snap,
    })

    print("[*] Building HTML report ...")
    _HTML_CONTENT = build_html(ca_info, results, domain, args.dc_ip,
                               diff_changes, prev_snap, curr_snap)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(_HTML_CONTENT)
        print(f"[+] Report saved to {args.output}")

    print(f"\n[+] Serving on http://0.0.0.0:{args.port}")
    print(f"    Open http://localhost:{args.port} in your browser")
    print(f"    Exports: /export/json  /export/excel  /export/history")
    print(f"    Ctrl+C to stop\n")

    try:
        HTTPServer(("0.0.0.0", args.port), Handler).serve_forever()
    except KeyboardInterrupt:
        print("\n[*] Stopped.")

if __name__ == "__main__":
    main()
